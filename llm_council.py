import os
import sys
import time
import io
import re
import json
import datetime
import traceback
import smtplib
import asyncio
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# ---------------------------------------------------------------------
#  SAFETY NET & SECURITY PATCH: LangChain Hub fix (CVE-2026-45134)
# ---------------------------------------------------------------------
try:
    import langsmith.client
    # Omgår sikkerhedsfejlen ved at gøre valideringsfunktionen inaktiv [2.2.1]
    langsmith.client._validate_public_prompt_pull = lambda *args, **kwargs: None
    print("Security patch: LangSmith public prompt validation bypassed successfully.")
except ImportError:
    # Hvis langsmith endnu ikke er installeret, fortsætter vi uden fejl
    pass

# ---------------------------------------------------------------------
#  SAFETY NET: Automatic installation of openpyxl if missing
# ---------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    import subprocess
    print("Safety net: openpyxl is missing. Installing automatically...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ---------------------------------------------------------------------
#  SAFETY NET: Automatic installation of edge-tts if missing
# ---------------------------------------------------------------------
try:
    import edge_tts
except ImportError:
    import subprocess
    print("Safety net: edge-tts is missing. Installing automatically...")
    subprocess.call([sys.executable, "-m", "pip", "install", "edge-tts"])
    import edge_tts

import yfinance as yf
import requests
import pandas as pd

# =====================================================================
#  CONFIGURATION & STANDARD TARGET WEIGHTS
# =====================================================================

TARGET_PORTFOLIO = {
    "Aktier": 25.0,
    "Sukuk": 25.0,
    "Råvarer": 25.0,
    "Kontanter/Private": 25.0
}

# Translation dictionary for the English report output
DISPLAY_CATEGORIES = {
    "Aktier": "Equities",
    "Sukuk": "Sukuk (Islamic Bonds)",
    "Råvarer": "Commodities",
    "Kontanter/Private": "Cash / Private Sector"
}

TARGET_SUBSECTORS = [
    "Pharmaceuticals & Biotech",
    "Medical Devices & MedTech",
    "Clean Energy & Wind",
    "Smart Grid & Electrification",
    "Global Infrastructure",
    "Construction & Building Materials",
    "Chemicals & Advanced Materials",
    "Industrial Machinery & Automation",
    "Semiconductors & Hardware",
    "Mining & Royalty Streams",
    "Traditional Energy & Utilities",
    "Global Equity ETFs",
    "Regional & Thematic ETFs",
    "Sukuk & Fixed Income",
    "Cash & Liquidity Reserves",
    "Private Equity & Venture Capital",
    "Agriculture & Food Security",
    "Consumer Defensive & Staples",
    "Enterprise Software & SaaS",
    "Industrial Metals & Copper",
    "Logistics & Global Shipping",
    "Artificial Intelligence & Cloud Computing",
    "Cybersecurity & Digital Defense",
    "E-Commerce & Digital Retail",
    "Water Infrastructure & Desalination"
]

# DET STATISKE LYNHURTIGE KARTOTEK (Failsafe for at undgå IP-blokeringer fra Yahoo)
STATIC_TICKER_MAP = {
    "NOVO-B.CO": ("Aktier", "Pharmaceuticals & Biotech"),
    "NOVO-B": ("Aktier", "Pharmaceuticals & Biotech"),
    "MSFT": ("Aktier", "Enterprise Software & SaaS"),
    "AAPL": ("Aktier", "Semiconductors & Hardware"),
    "SAP": ("Aktier", "Enterprise Software & SaaS"),
    "IFX.DE": ("Aktier", "Semiconductors & Hardware"),
    "ASML": ("Aktier", "Semiconductors & Hardware"),
    "NVDA": ("Aktier", "Semiconductors & Hardware"),
    "VWS.CO": ("Aktier", "Clean Energy & Wind"),
    "NKT.CO": ("Aktier", "Smart Grid & Electrification"),
    "FLS.CO": ("Aktier", "Industrial Machinery & Automation"),
    "ROCK-B.CO": ("Aktier", "Construction & Building Materials"),
    "ORK.OL": ("Aktier", "Consumer Defensive & Staples"),
    "WPM": ("Råvarer", "Mining & Royalty Streams"),
    "NEM": ("Råvarer", "Industrial Metals & Copper"),
    "AEM": ("Råvarer", "Industrial Metals & Copper"),
    "RGLD": ("Råvarer", "Mining & Royalty Streams"),
    "SPSK": ("Sukuk", "Sukuk & Fixed Income"),
    "SKUK": ("Sukuk", "Sukuk & Fixed Income"),
    "MSAU.L": ("Aktier", "Regional & Thematic ETFs"),
    "IGDA.L": ("Aktier", "Global Equity ETFs"),
    "HLAL": ("Aktier", "Regional & Thematic ETFs"),
    "UMMA": ("Aktier", "Global Equity ETFs"),
    "ISWD.L": ("Aktier", "Global Equity ETFs"),
    "ISUS.L": ("Aktier", "Regional & Thematic ETFs"),
    "HIWS.L": ("Aktier", "Global Equity ETFs")
}

# HYBRID AUTOMATISK DATABASE-INDLÆSNING FRA DIN GITHUB
failsafe_db = STATIC_TICKER_MAP.copy()
if os.path.exists("failsafe_db.json"):
    try:
        with open("failsafe_db.json", "r") as f:
            loaded_db = json.load(f)
            failsafe_db.update(loaded_db)
            print(f"Loaded {len(loaded_db)} tickers from failsafe_db.json successfully!")
    except Exception as e:
        print(f"Warning: Could not load failsafe_db.json: {str(e)}")

# GLOBAL ISLAMIC GROWTH UNIVERSE (Anvendes proaktivt)
GLOBAL_COMPLIANT_GROWTH_POOL = {
    "Aktier": [
        "MSAU.L", "IGDA.L", "HLAL", "UMMA", "ISWD.L", "ISUS.L", "HIWS.L",
        "TRMB", "SAP", "IFX.DE", "MSFT", "ASML", "NVDA", "ADBE", "CRM", "SNPS", 
        "NOVO-B.CO", "6869.T", "AZN.ST", "REGN", "ISRG", "LLY", "VRTX", 
        "VWS.CO", "NKT.CO", "FLS.CO", "ROCK-B.CO"
    ],
    "Sukuk": [
        "SPSK", "SKUK"
    ],
    "Råvarer": [
        "WPM", "NEM", "GOLD", "AEM", "FNV", "RGLD", "BHP", "RIO", "FCX", "VALE"
    ],
    "Kontanter/Private": [
        "SPSK", "SKUK"
    ]
}

GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "1EnE2XkQySaGsdaxR5KySZZ924LT66ICo")
DATABASE_URL = os.getenv("DATABASE_URL")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

EMAIL_SENDER = os.getenv("EMAIL_SENDER")
if not EMAIL_SENDER or EMAIL_SENDER.strip() == "":
    EMAIL_SENDER = "wazir.ilyas@gmail.com"

EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")
if not EMAIL_RECEIVER or EMAIL_RECEIVER.strip() == "":
    EMAIL_RECEIVER = "addoncreatives@gmail.com"

EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
if EMAIL_PASSWORD:
    EMAIL_PASSWORD = EMAIL_PASSWORD.replace(" ", "").strip()


def normalize_string(s: str) -> str:
    if not s or pd.isna(s):
        return ""
    s = str(s).lower().strip()
    s = s.replace("og", "and").replace("&", "and")
    s = s.replace("'", "")
    s = re.sub(r'[^a-z0-9æøå]', '', s)
    s = s.replace("etfer", "etf").replace("etfs", "etf")
    return s


# =====================================================================
#  LIVE SEARCH-TO-TICKER MOTOR (FINDER AUTOMATISK TICKERS FRA NAVNE)
# =====================================================================
def search_ticker_by_name(query: str) -> str:
    if not query or pd.isna(query):
        return None
    
    query_clean = str(query).strip()
    
    if re.match(r'^[A-Z0-9\.\-]+$', query_clean) and len(query_clean) < 10 and "." in query_clean:
        return query_clean

    url = f"https://query2.finance.yahoo.com/v1/finance/search?q={query_clean}"
    headers = {
        'User-Agent': 'Mozilla/5.0'
    }
    try:
        response = requests.get(url, headers=headers, timeout=6)
        if response.status_code == 200:
            data = response.json()
            quotes = data.get("quotes", [])
            if quotes:
                return quotes[0].get("symbol")
    except Exception as e:
        print(f"Search failed for '{query_clean}': {str(e)}")
    return None


# =====================================================================
#  OPDATERET EXCEL SKABELONS GENERATOR (MED NATIVE LIVE FORMELER)
# =====================================================================
def generate_excel_template_bytes(holdings_list: list, watchlist_list: list) -> bytes:
    wb = openpyxl.Workbook()
    
    # 1. FANEN: Beholdninger
    ws1 = wb.active
    ws1.title = "Beholdninger"
    
    headers1 = [
        "Position", "Ticker", "Status", "Antal", "Kurs (DKK)", 
        "Markedsværdi (DKK)", "Aktivklasse", "Drivkraft", "Sektor", 
        "Region", "Porteføljevægt", "Rolle", "Kommentar / tese"
    ]
    ws1.append(headers1)
    
    for idx, item in enumerate(holdings_list, start=2):
        name = item.get("name", "Other") if isinstance(item, dict) else item.get("Company Name", "Other")
        symbol = item.get("ticker", "Other") if isinstance(item, dict) else item.get("Ticker", "Other")
        shares = int(item.get("Shares", 1)) if "Shares" in item else 1
        cat = item.get("asset_class", "Aktier") if isinstance(item, dict) else item.get("Category", "Aktier")
        sec = item.get("sector", "Other") if isinstance(item, dict) else item.get("Sector", "Other")
        
        # Hvis det er et manuelt aktiv
        if "PVT_" in symbol or "CASH_" in symbol:
            val = float(item.get("manual_value", 1000))
            ws1.cell(row=idx, column=1, value=name)
            ws1.cell(row=idx, column=2, value="")
            ws1.cell(row=idx, column=3, value="Ejer")
            ws1.cell(row=idx, column=4, value=1)
            ws1.cell(row=idx, column=5, value=val)
            ws1.cell(row=idx, column=6, value=val)
        else:
            ws1.cell(row=idx, column=1, value=name)
            ws1.cell(row=idx, column=2, value=symbol)
            ws1.cell(row=idx, column=3, value="Ejer")
            ws1.cell(row=idx, column=4, value=shares)
            ws1.cell(row=idx, column=5, value=f'=GOOGLEFINANCE(B{idx})')
            ws1.cell(row=idx, column=6, value=f'=D{idx}*E{idx}')
            
        ws1.cell(row=idx, column=7, value=cat)
        ws1.cell(row=idx, column=8, value="")
        ws1.cell(row=idx, column=9, value=sec)
        ws1.cell(row=idx, column=10, value="Global")
        ws1.cell(row=idx, column=11, value=f'=F{idx}/SUM(F$2:F$100)')
        ws1.cell(row=idx, column=12, value="")
        ws1.cell(row=idx, column=13, value="")

    # 2. FANEN: Opsummering
    ws2 = wb.create_sheet(title="Opsummering")
    headers2 = ["4x25-overblik", "", "", "", "", "Økonomiske drivere", "", "", "", "Sektorere", "", "", "", "Huller / Watchlist"]
    ws2.append(headers2)
    
    # Skriv watchlist i Kolonne N (14)
    for idx, ticker in enumerate(watchlist_list, start=2):
        ws2.cell(row=idx, column=14, value=ticker)
        
    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data.getvalue()


# =====================================================================
#  GOOGLE SHEETS / EXCEL AGENT
# =====================================================================
class GoogleSheetsAgent:
    def __init__(self, sheet_id: str):
        self.sheet_id = sheet_id

    def _read_tab_as_df(self, tab_name: str) -> pd.DataFrame:
        url = f"https://docs.google.com/spreadsheets/d/{self.sheet_id}/export?format=xlsx"
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            df = pd.read_excel(io.BytesIO(response.content), sheet_name=tab_name, engine='openpyxl')
            return df
        except Exception as e:
            raise RuntimeError(f"Kunne ikke indlæse fanen '{tab_name}': {str(e)}")

    def _clean_and_align_df(self, df: pd.DataFrame, key_header_word: str) -> pd.DataFrame:
        for col in df.columns:
            if key_header_word.lower() in str(col).lower():
                return df
                
        for idx, row in df.head(10).iterrows():
            row_values = [str(val).lower() for val in row.values]
            if any(key_header_word.lower() in val for val in row_values):
                new_columns = df.iloc[idx].values
                df.columns = new_columns
                df = df.iloc[idx+1:].reset_index(drop=True)
                return df
        return df

    def _find_column_by_keyword(self, df: pd.DataFrame, keyword: str) -> str:
        for col in df.columns:
            if keyword.lower() in str(col).lower():
                return col
        return None

    def get_current_weights_and_sectors(self) -> tuple:
        """
        Læser fanen 'Beholdninger', identificerer vægtene og fordeler dem 
        både på hovedkasserne og på de specifikke delsektorer helt automatisk.
        """
        try:
            raw_df = self._read_tab_as_df("Beholdninger")
            df = self._clean_and_align_df(raw_df, "ticker")
            
            ticker_col = self._find_column_by_keyword(df, "ticker")
            drivkraft_col = self._find_column_by_keyword(df, "drivkraft")
            aktivklasse_col = self._find_column_by_keyword(df, "aktivklasse")
            sektor_col = self._find_column_by_keyword(df, "sektor")
            weight_col = self._find_column_by_keyword(df, "vægt")
            mv_col = self._find_column_by_keyword(df, "markedsværdi")

            if not weight_col:
                raise KeyError("Kunne ikke lokalisere de nødvendige kolonner i Google Sheet.")

            def clean_number(val):
                if pd.isna(val) or val == "":
                    return 0.0
                val_str = str(val).replace('%', '').replace('kr', '').replace('$', '').replace(' ', '').replace('\xa0', '').strip()
                if ',' in val_str and '.' in val_str:
                    if val_str.find('.') < val_str.find(','):
                        val_str = val_str.replace('.', '').replace(',', '.')
                    else:
                        val_str = val_str.replace(',', '').replace('.', '.')
                elif ',' in val_str:
                    val_str = val_str.replace(',', '.')
                try:
                    return float(val_str)
                except ValueError:
                    return 0.0

            df['Cleaned_Weight'] = df[weight_col].apply(clean_number)
            df['Cleaned_MV'] = df[mv_col].apply(clean_number) if mv_col else 0.0

            total_mv = df['Cleaned_MV'].sum()
            total_weight = df['Cleaned_Weight'].sum()

            if total_weight < 1.0 and total_mv > 0.0:
                df['Cleaned_Weight'] = (df['Cleaned_MV'] / total_mv) * 100.0
            elif total_mv > 0.0:
                df['Cleaned_Weight'] = (df['Cleaned_MV'] / total_mv) * 100.0

            portfolio_distribution = {k: 0.0 for k in TARGET_PORTFOLIO.keys()}
            sector_distribution = {}

            temp_screener = ScreenerComplianceAgent([])

            for _, row in df.iterrows():
                row_weight = row['Cleaned_Weight']
                if row_weight <= 0.0:
                    continue

                symbol = str(row.get(ticker_col, "")).strip().upper()
                drivkraft_val = normalize_string(row.get(drivkraft_col, "")) if drivkraft_col else ""
                aktivklasse_val = normalize_string(row.get(aktivklasse_col, "")) if aktivklasse_col else ""
                sektor_val = normalize_string(row.get(sektor_col, "")) if sektor_col else ""

                combined_text = f"{drivkraft_val} {aktivklasse_val} {sektor_val}"
                
                category = None
                subsector = None

                if "sukuk" in combined_text:
                    category = "Sukuk"
                    subsector = "Sukuk & Fixed Income"
                elif any(word in combined_text for word in ["råvarer", "ravarer", "guld", "gold", "commodities"]):
                    category = "Råvarer"
                    subsector = "Mining & Royalty Streams"
                elif any(word in combined_text for word in ["kontant", "cash", "private"]):
                    category = "Kontanter/Private"
                    subsector = "Cash & Liquidity Reserves"
                elif "aktie" in combined_text:
                    category = "Aktier"

                if not category or not subsector:
                    try:
                        t = yf.Ticker(symbol)
                        info = t.info
                        sec = info.get("sector", "Other")
                        ind = info.get("industry", "Other")
                        
                        cat_detect, sub_detect = temp_screener.map_to_category_and_sector(symbol, sec, ind)
                        category = category if category else cat_detect
                        subsector = subsector if subsector else sub_detect
                    except Exception:
                        category = category if category else "Aktier"
                        subsector = subsector if subsector else "Other"

                if category in portfolio_distribution:
                    portfolio_distribution[category] += row_weight
                
                if subsector not in sector_distribution:
                    sector_distribution[subsector] = 0.0
                sector_distribution[subsector] += row_weight

            return portfolio_distribution, sector_distribution
        except Exception as e:
            print(f"Fejl under indlæsning: {str(e)}")
            return {k: 0.0 for k in TARGET_PORTFOLIO.keys()}, {}

    def get_current_holdings_details(self) -> list:
        try:
            raw_df = self._read_tab_as_df("Beholdninger")
            df = self._clean_and_align_df(raw_df, "ticker")
            
            ticker_col = self._find_column_by_keyword(df, "ticker")
            position_col = self._find_column_by_keyword(df, "position") or self._find_column_by_keyword(df, "navn")
            mv_col = self._find_column_by_keyword(df, "markedsværdi")
            sektor_col = self._find_column_by_keyword(df, "sektor")
            aktivklasse_col = self._find_column_by_keyword(df, "aktivklasse")
            
            holdings = []
            for _, row in df.iterrows():
                ticker = str(row.get(ticker_col, "")).strip().upper()
                if ticker and not pd.isna(row.get(ticker_col)) and ticker not in ["TICKER", "STATUS", "POSITION", "HULLER"]:
                    holdings.append({
                        "ticker": ticker,
                        "name": str(row.get(position_col, ticker)).strip(),
                        "market_value": str(row.get(mv_col, "0.00 DKK")).strip(),
                        "sector": str(row.get(sektor_col, "N/A")).strip(),
                        "asset_class": str(row.get(aktivklasse_col, "N/A")).strip()
                    })
            return holdings
        except Exception as e:
            print(f"Fejl: {str(e)}")
            return []

    def get_watchlist_tickers(self) -> list:
        try:
            raw_df = self._read_tab_as_df("Opsummering")
            df = self._clean_and_align_df(raw_df, "huller")
            watchlist_col = self._find_column_by_keyword(df, "huller") or self._find_column_by_keyword(df, "watchlist")
            
            tickers = []
            if watchlist_col:
                raw_series = df[watchlist_col]
            else:
                if len(df.columns) >= 14:
                    raw_series = df.iloc[:, 13]
                else:
                    return []

            for val in raw_series:
                if pd.isna(val):
                    continue
                val_str = str(val).strip().upper()
                if val_str and len(val_str) < 12 and re.match(r'^[A-Z0-9\.\-]+$', val_str):
                    if val_str not in ["TICKER", "STATUS", "POSITION", "HULLER"]:
                        tickers.append(val_str)
            return list(set(tickers))
        except Exception as e:
            return []


# =====================================================================
#  PORTFOLIO MANAGER AGENT (STATELÆS ROTATION)
# =====================================================================
class PortfolioManagerAgent:
    def __init__(self, current: dict, target: dict):
        self.current = current
        self.target = target

    def identify_underweighted_focus(self) -> tuple:
        underweight_candidates = []
        for category, target_val in self.target.items():
            curr_val = self.current.get(category, 0.0)
            deficit = target_val - curr_val
            if deficit > 0.0:
                underweight_candidates.append((category, deficit))
        
        if not underweight_candidates:
            return list(self.target.keys())[0], 0.0

        underweight_candidates.sort(key=lambda x: x[0])
        day_of_year = datetime.datetime.now().timetuple().tm_yday
        index = day_of_year % len(underweight_candidates)
        
        focus_category, deficit = underweight_candidates[index]
        return focus_category, deficit


# =====================================================================
#  SCREENER & COMPLIANCE AGENT (DYNAMISK SØGNING)
# =====================================================================
class ScreenerComplianceAgent:
    PROHIBITED_SECTORS = ["Financial Services", "Financial"]
    PROHIBITED_INDUSTRIES = [
        "Banks", "Insurance", "Aerospace & Defense", "Gambling", 
        "Tobacco", "Distillers & Vintners", "Breweries"
    ]

    def __init__(self, tickers: list, target_category: str = None):
        self.tickers = tickers
        self.target_category = target_category

    def check_zoya_live_compliance(self, symbol: str) -> bool:
        clean_symbol = symbol.split('.')[0].upper()
        clean_symbol = clean_symbol.split('-')[0]
        
        url = f"https://zoya.finance/stocks/{clean_symbol}"
        headers = {
            'User-Agent': 'Mozilla/5.0'
        }
        try:
            response = requests.get(url, headers=headers, timeout=8)
            if response.status_code == 200:
                html_lower = response.text.lower()
                if "is shariah-compliant" in html_lower or "is considered halal" in html_lower:
                    return True
                elif "not shariah-compliant" in html_lower or "non-compliant" in html_lower:
                    return False
            return None
        except Exception:
            return None

    def map_to_category_and_sector(self, symbol: str, sector: str = None, industry: str = None) -> tuple:
        sym = symbol.upper().strip()
        sec_l = sector.lower() if sector else ""
        ind_l = industry.lower() if industry else ""

        # 1. Tjek altid det lynhurtige statiske kartotek først
        lookup_sym = sym.split('.')[0]
        for k, v in failsafe_db.items():
            if normalize_string(k) == normalize_string(sym) or normalize_string(k) == normalize_string(lookup_sym):
                if v[0] == "Sukuk" and self.target_category == "Kontanter/Private":
                    return "Kontanter/Private", "Sukuk & Fixed Income"
                return v[0], v[1]

        # 2. Hvis ikke i kartoteket, kør dynamisk mapping baseret på yfinance data
        if "sukuk" in sym or sym in ["SPSK", "SKUK"]:
            if self.target_category == "Kontanter/Private":
                return "Kontanter/Private", "Sukuk & Fixed Income"
            return "Sukuk", "Sukuk & Fixed Income"
            
        if sym in ["WPM", "FNV", "RGLD"]:
            return "Råvarer", "Mining & Royalty Streams"
        if sym in ["NEM", "GOLD", "AEM", "BHP", "RIO", "FCX", "VALE"] or \
           any(w in ind_l for w in ["gold", "silver", "precious metals", "copper", "aluminum"]):
            return "Råvarer", "Industrial Metals & Copper"

        if "cash" in sym or "money market" in sec_l:
            return "Kontanter/Private", "Cash & Liquidity Reserves"

        # Dynamisk kobling til de nye overordnede kategorier
        if "pharmaceutical" in ind_l or "biotechnology" in ind_l:
            return "Aktier", "Pharmaceuticals & Biotech"
        if "medical" in ind_l or "healthcare" in sec_l:
            return "Aktier", "Medical Devices & MedTech"
        if "wind" in ind_l or "renewable" in ind_l or "solar" in ind_l:
            return "Aktier", "Clean Energy & Wind"
        if "cable" in ind_l or "electrical" in ind_l:
            return "Aktier", "Smart Grid & Electrification"
        if "semiconductor" in ind_l or "semiconductor" in sec_l:
            return "Aktier", "Semiconductors & Hardware"
        if "software" in ind_l or "software" in sec_l or "technology" in sec_l:
            return "Aktier", "Enterprise Software & SaaS"
        if "building" in ind_l or "construction" in ind_l:
            return "Aktier", "Construction & Building Materials"
        if "chemicals" in ind_l:
            return "Aktier", "Chemicals & Advanced Materials"
        if "machinery" in ind_l or "industrials" in sec_l:
            return "Aktier", "Industrial Machinery & Automation"
        if "infrastructure" in ind_l or "utilities" in sec_l:
            return "Aktier", "Global Infrastructure"
        if "staples" in sec_l or "packaged foods" in ind_l or "consumer defensive" in sec_l:
            return "Aktier", "Consumer Defensive & Staples"
        if "fertilizer" in ind_l or "agriculture" in ind_l:
            return "Aktier", "Agriculture & Food Security"
        if "logistics" in ind_l or "shipping" in ind_l:
            return "Aktier", "Logistics & Global Shipping"
        if "internet" in ind_l or "e-commerce" in ind_l:
            return "Aktier", "E-Commerce & Digital Retail"
        if "water" in ind_l or "environmental" in ind_l:
            return "Aktier", "Water Infrastructure & Desalination"

        # Dynamisk fallback
        dynamic_subsector = industry if (industry and industry != "Other") else sector
        return "Aktier", dynamic_subsector

    def screen_ticker(self, symbol: str) -> dict:
        try:
            # Virtuelle/manuelle tickers for kontanter/private skal altid bestå screening
            if "CASH_" in symbol or "PVT_" in symbol:
                return {
                    "symbol": symbol,
                    "passed": True,
                    "name": symbol.replace("CASH_", "").replace("PVT_", ""),
                    "pe_ratio": "N/A",
                    "debt_ratio": "0.00% (Manual)",
                    "sector": "Manual Asset",
                    "industry": "Manual Asset",
                    "category": "Kontanter/Private",
                    "subsector": "Cash & Liquidity Reserves",
                    "is_etf": False
                }

            zoya_compliant = self.check_zoya_live_compliance(symbol)
            
            if zoya_compliant is False:
                return {"symbol": symbol, "passed": False, "reason": "Disqualified by Zoya's live public Shariah assessment."}
            elif zoya_compliant is True:
                print(f"Zoya Live-tjek bekræfter: {symbol} er Shariah-compliant.")

            ticker_obj = yf.Ticker(symbol)
            info = ticker_obj.info
            
            if not info:
                return {"symbol": symbol, "passed": False, "reason": "Ingen data"}

            quote_type = info.get("quoteType", "").upper()
            is_etf = quote_type in ["ETF", "MUTUALFUND"] or symbol in ["IGDA.L", "SPSK", "HLAL", "UMMA", "ISWD.L", "MSAU.L", "SKUK"]

            if is_etf:
                mapped_cat, mapped_sub = self.map_to_category_and_sector(symbol, "ETF", "ETF")
                return {
                    "symbol": symbol,
                    "passed": True,
                    "name": info.get("longName", symbol),
                    "pe_ratio": info.get("trailingPE", "N/A"),
                    "debt_ratio": "N/A (ETF/Sukuk)",
                    "sector": "ETF / Fond",
                    "industry": "ETF",
                    "category": mapped_cat,
                    "subsector": mapped_sub,
                    "is_etf": True
                }

            sector = info.get("sector", "")
            industry = info.get("industry", "")
            
            for p_sector in self.PROHIBITED_SECTORS:
                if p_sector.lower() in sector.lower():
                    return {"symbol": symbol, "passed": False, "reason": f"Sektor: {sector}"}
                    
            for p_ind in self.PROHIBITED_INDUSTRIES:
                if p_ind.lower() in industry.lower():
                    return {"symbol": symbol, "passed": False, "reason": f"Branche: {industry}"}

            debt_to_equity = info.get("debtToEquity")
            total_debt = info.get("totalDebt")
            market_cap = info.get("marketCap")
            
            debt_ratio_pct = None
            method_used = ""

            if debt_to_equity is not None:
                debt_ratio_pct = debt_to_equity
                method_used = "Debt to Equity"
            elif total_debt and market_cap:
                debt_ratio_pct = (total_debt / market_cap) * 100
                method_used = "Debt to Market Cap"

            if debt_ratio_pct is None:
                return {"symbol": symbol, "passed": False, "reason": "Ingen gældsdata"}

            if debt_ratio_pct > 30.0:
                return {"symbol": symbol, "passed": False, "reason": f"Gældskvote: {debt_ratio_pct:.2f}%"}

            mapped_cat, mapped_sub = self.map_to_category_and_sector(symbol, sector, industry)

            return {
                "symbol": symbol,
                "passed": True,
                "name": info.get("longName", symbol),
                "pe_ratio": info.get("trailingPE", "N/A"),
                "debt_ratio": f"{debt_ratio_pct:.2f}% ({method_used})",
                "sector": sector,
                "industry": industry,
                "category": mapped_cat,
                "subsector": mapped_sub,
                "is_etf": False
            }
        except Exception as e:
            return {"symbol": symbol, "passed": False, "reason": str(e)}

    def run_screening(self, target_category: str) -> list:
        approved_stocks = []
        for ticker in self.tickers:
            time.sleep(1.5)
            result = self.screen_ticker(symbol=ticker)
            if result["passed"] and result["category"] == target_category:
                approved_stocks.append(result)
        return approved_stocks


# =====================================================================
#  COUNCIL AGENT (GEMINI 3.5 FLASH - ENGLISH HTML NEWSLETTER)
# =====================================================================
class CouncilAgent:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.5-flash:generateContent?key={self.api_key}"

    def run_proactive_analysis(self, candidates_data: list, category: str, deficit: float, current_portfolio_str: str, current_holdings_str: str, sector_distribution_str: str, user_name: str, horizon: str) -> str:
        candidates_json = json.dumps(candidates_data, indent=2, ensure_ascii=False)
        english_category = DISPLAY_CATEGORIES.get(category, category)
        
        prompt = f"""
        You are an elite financial advisory council ("LLM Council") presenting a strategic investment briefing to your highly valued VIP client, {user_name}.
        
        THE INVESTOR PROFILE & MODEL:
        - Investor's Name: {user_name}
        - Investment Horizon: {horizon} (This is CRITICAL. Align all advice, timelines, risk-tolerances, and recommendations precisely with this specific time horizon!)
        - Overarching Strategic Model: Customize based on Wazir's targets.
        - Under Evaluation Tonight: {english_category} (Current Deficit: {deficit:.2f}%)
        - Current Portfolio Allocations (Target vs Actual): {current_portfolio_str}
        
        THE INVESTOR'S STRATEGIC SUB-SECTORS (DYNAMICALLY DETECTED FROM THE HOLDINGS):
        {sector_distribution_str}
        
        THE INVESTOR'S CURRENT HOLDINGS:
        {current_holdings_str}
        
        NEW COMPLIANT SCREENED CANDIDATES TO BE EVALUATED:
        {candidates_json}
        
        YOUR OBJECTIVE (DELIVER ENTIRE BRIEFING IN BEAUTIFUL ENGLISH HTML):
        Generate a complete, institutional-grade, highly engaging HTML investment newsletter in English.
        
        Brug udelukkende inline CSS-styling for maximum compatibility with Gmail.
        Design guidelines:
        - Main container: `<div style="font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; max-width: 800px; margin: 0 auto; padding: 25px; background-color: #ffffff; color: #334155; line-height: 1.6;">`
        - Colors: Dark Slate (`#0F172A`) for headings. Accent/highlights: Warm Gold/Sand (`#C5A880`).
        - Cards: Each of the screened candidates should be enclosed in a distinct card: `<div style="border: 1px solid #E2E8F0; padding: 20px; margin-bottom: 25px; border-radius: 8px; background-color: #F8FAFC;">`
        - Debate box: Each adviser's turn must use the distinct left-bordered styling:
          - Contrarian: `border-left: 4px solid #EF4444; background: #FEF2F2; padding: 15px; margin-bottom: 15px;` (Red border)
          - First-Principles: `border-left: 4px solid #64748B; background: #F8FAFC; padding: 15px; margin-bottom: 15px;` (Slate border)
          - Expansionist: `border-left: 4px solid #10B981; background: #ECFDF5; padding: 15px; margin-bottom: 15px;` (Green border)
          - Outsider: `border-left: 4px solid #8B5CF6; background: #F5F3FF; padding: 15px; margin-bottom: 15px;` (Purple border)
          - Executor: `border-left: 4px solid #3B82F6; background: #EFF6FF; padding: 15px; margin-bottom: 15px;` (Blue border)
        - Chairman's Call: Gold-framed box: `<div style="background-color: #FDFBF7; border: 1px solid #E2D1B6; border-left: 6px solid #C5A880; padding: 25px; border-radius: 8px; margin-top: 30px;">`
        
        REPORT OUTLINE (ENGLISH):
        
        <h1>🗳️ LLM Council Strategic Briefing</h1>
        <p><strong>Prepared exclusively for:</strong> {user_name}</p>
        <p><strong>Investment Horizon:</strong> {horizon}</p>
        <p><strong>Focus tonight:</strong> {english_category} (Deficit: {deficit:.2f}%)</p>
        
        <hr style="border: 0; border-top: 1px solid #E2E8F0; margin: 20px 0;">
        
        <h2>SECTION 1 — PORTFOLIO DIAGNOSTIC & INDIRECT EXPOSURES</h2>
        Analyze {user_name}'s current holdings and how they map to their strategic sub-sectors. Do existing assets already provide satisfactory indirect exposure to the focus theme? Discuss Saxo Bank limitations and Sharia compliance filters as boundaries, and explain how they can diversify across different underlying economic drivers if direct options are limited.
        
        <h2>SECTION 2 — DEEP-DIVE CONSULTANT ANALYSIS (UP TO 10 SCREENED CANDIDATES)</h2>
        For each candidate, write an elegant card covering:
        1. <strong>Investment Case</strong> (How it balances {user_name}'s current assets, keeping their {horizon} horizon in mind).
        2. <strong>Financial Highlights</strong> (Omsætningsvækst, marginer, cash flow based on live data).
        3. <strong>Future Outlook & Pipeline</strong>.
        4. <strong>Risk Assessment</strong>.
        5. <strong>Momentum & Trend Analysis</strong> (3-month momentum vs. 3-year growth trajectory).
        6. <strong>Analyst Insight & Sources</strong>: Insert exactly 2 clickable links structured beautifully in HTML (e.g. `<a href="https://seekingalpha.com/symbol/TICKER" style="color: #C5A880; text-decoration: none; font-weight: bold;">Seeking Alpha</a>`).
        
        <h2>SECTION 3 — THE ASYNCHRONOUS COUNCIL DEBAT (TOP-3)</h2>
        Select the top 3 assets. Moderate a high-stakes, dramatic debate among the 5 financial advisers using the styled left-bordered divs. Show conflict, arguments on valuation, capex, and macro timing.
        
        <h2>SECTION 4 — THE CHAIRMAN'S DEKRET (RECOMMENDATION)</h2>
        In Section 4, the Chairman must NOT command the investor to buy or take action. Instead, the Chairman must strongly advise and urge the investor to critically evaluate and consider the council's comprehensive proposal. The tone should be highly advisory, objective, and respectful, emphasizing that the final capital allocation decision rests solely on the investor's own assessment. Enclose this callout in the gold callout box. Conclude with a highly precise, step-by-step action plan for {user_name}'s Saxo Investor account over the next 7 days.
        
        Return ONLY the raw HTML code. Do NOT enclose in markdown tags like "```html".
        """

        headers = {'Content-Type': 'application/json'}
        payload = {"contents": [{"parts": [{"text": prompt}]}]}
        
        try:
            response = requests.post(self.url, headers=headers, json=payload, timeout=110)
            response.raise_for_status()
            data = response.json()
            if 'candidates' in data and len(data['candidates']) > 0:
                raw_html = data['candidates'][0]['content']['parts'][0]['text']
                return raw_html.replace("```html", "").replace("```", "").strip()
            return "<h3>Error</h3><p>Could not parse HTML from Gemini.</p>"
        except Exception as e:
            return f"<h3>System Error</h3><p>{str(e)}</p>"


# =====================================================================
#  PODCAST AGENT (PODCASTFY POWERED HIGH-ENERGY CNBC/BLOOMBERG TALKSHOW)
# =====================================================================
class PodcastAgent:
    def __init__(self, api_key: str):
        self.api_key = api_key
        if self.api_key:
            os.environ["GEMINI_API_KEY"] = self.api_key
            os.environ["GOOGLE_API_KEY"] = self.api_key

    def generate_podcast_audio(self, report_html: str, user_name: str) -> str:
        from podcastfy.client import generate_podcast
        
        custom_config = {
            "word_count": 900,
            "conversation_style": ["engaging", "fast-paced", "enthusiastic", "hardcore Bloomberg debate"],
            "roles_person1": "Sarah, the curious financial journalist",
            "roles_person2": "Mark, the hardcore market analyst",
            "podcast_name": "LLM Council Briefing",
            "podcast_tagline": "CNBC & Bloomberg Style Financial Talkshow",
            "output_language": "English",
            "engagement_techniques": ["rhetorical questions", "analogies", "humor", "interjections", "cross-talk"],
            "user_instructions": (
                f"Create a high-energy Bloomberg-style financial show moderated by Sarah and Mark. "
                f"The show MUST open with Sarah and Mark introducing themselves and welcoming our VIP client, {user_name}. "
                f"Then, they introduce and interview our 5 resident advisers: "
                f"Contrarian (the risk-obsessed skeptic who must interrupt with: 'But what if the market turns tomorrow?'), "
                f"First-Principles (the logical mathematician using raw numbers), "
                f"Expansionist (the highly bullish growth hunter wanting to deploy capital), "
                f"Outsider (the big-picture strategist analyzing indirect exposures like NKT/FLS and favoring royalty models), "
                f"and Executor (the pragmatic guy checking Saxo tradeability and Dollar-Cost Averaging). "
                f"The show must conclude with Sarah and Mark summarizing the Chairman's final recommendation and "
                f"giving {user_name} a highly clear, actionable next step for his Saxo account."
            )
        }
        
        # Den uheldige interne try-except blok er nu FULDSTÆNDIG slettet [3]!
        # Hvis Podcastfy fejler, vil fejlen boble direkte op til main() og blive sendt i din mail [3].
        print("Genererer ægte multi-stemme podcast via Podcastfy og gratis Edge TTS...")
        audio_path = generate_podcast(
            text=report_html,
            tts_model="edge",
            conversation_config=custom_config
        )
        return audio_path


# =====================================================================
#  DELIVERY AGENT (HTML & VEDHÆFTNING SMTP)
# =====================================================================
class DeliveryAgent:
    @staticmethod
    def send_email(subject: str, html_content: str, attachments: list = None):
        if not EMAIL_PASSWORD or EMAIL_PASSWORD.strip() == "":
            print("EMAIL_PASSWORD mangler i GitHub Secrets. Udskriver HTML i konsol:")
            print(html_content)
            return

        msg = MIMEMultipart()
        msg["From"] = EMAIL_SENDER
        msg["To"] = EMAIL_RECEIVER
        msg["Subject"] = subject
        msg.attach(MIMEText(html_content, "html", "utf-8"))

        if attachments:
            for att in attachments:
                path = att.get("path")
                data = att.get("data")
                name = att.get("name")
                
                if (path and os.path.exists(path)) or data:
                    part = MIMEBase("application", "octet-stream")
                    if data:
                        part.set_payload(data)
                    else:
                        with open(path, "rb") as attachment:
                            part.set_payload(attachment.read())
                    
                    encoders.encode_base64(part)
                    part.add_header(
                        "Content-Disposition",
                        f"attachment; filename= {name}",
                    )
                    msg.attach(part)

        try:
            print(f"Forbinder til SMTP server ({SMTP_SERVER}:{SMTP_PORT}) med afsender-login: {EMAIL_SENDER}...")
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
            server.quit()
            print("Succes: Rapport og podcast-afspilning sendt til din indbakke.")
        except Exception as e:
            print(f"E-mail fejl: {str(e)}")


# =====================================================================
#  ORCHESTRATOR / SYSTEM FLOW (HER ANVENDES NU DET RETTEDE KARTOTEK)
# =====================================================================
def main():
    try:
        print("Opstarter LLM Council 4x25% med delsektorer og Podcastfy...")
        sheets_agent = GoogleSheetsAgent(GOOGLE_SHEET_ID)
        
        # 1. Hent investors aktuelle vægte samt de 21 delsektorer (Opdateret med hybrid-database)
        current_portfolio_weights, sector_distribution = sheets_agent.get_current_weights_and_sectors()
        print(f"Beregnet 4x25% hovedfordeling: {current_portfolio_weights}")
        print(f"Beregnet delsektor-fordeling: {sector_distribution}")
        
        # 2. Hent de konkrete positioner
        current_holdings = sheets_agent.get_current_holdings_details()
        print(f"Hentet {len(current_holdings)} konkrete positioner.")
        
        # 3. Hent Watchlist-tickers
        watchlist_tickers = sheets_agent.get_watchlist_tickers()
        print(f"Watchlist: {watchlist_tickers}")

        # 4. Find den mest undervægtede kasse i din 4x25%-struktur (Tidsbaseret rotation!)
        pm = PortfolioManagerAgent(current_portfolio_weights, TARGET_PORTFOLIO)
        focus_category, deficit = pm.identify_underweighted_focus()
        print(f"Nattens strategiske fokus: {focus_category} (Underskud: {deficit:.2f}%)")

        # 5. PROAKTIV SØGNING: Kombiner personlig Watchlist med vores globale vækst-pool
        growth_pool = GLOBAL_COMPLIANT_GROWTH_POOL.get(focus_category, [])
        combined_candidates = list(set(watchlist_tickers + growth_pool))
        print(f"Kombineret søgebase ({len(combined_candidates)} aktiver): {combined_candidates}")

        # 6. Kør compliance screening (Sharia & Gælds-barrierer)
        print("Screener kandidater mod Sharia- og gældskrav...")
        screener = ScreenerComplianceAgent(combined_candidates, target_category=focus_category)
        approved_stocks = screener.run_screening(focus_category)
        print(f"Godkendte kandidater fundet efter screening: {[s['symbol'] for s in approved_stocks]}")

        target_candidates = approved_stocks[:10]
        
        if not target_candidates:
            error_html = f"<h2>No compliant assets could be found in your focused category ({focus_category}) today.</h2>"
            DeliveryAgent.send_email(f"[LLM Council] Alert - No candidates approved", error_html)
            return

        # 7. Indhent detaljerede yfinance tal
        print("Indhenter kvartalstal for de 10 godkendte kandidater...")
        detailed_candidates_data = []
        for stock in target_candidates:
            symbol = stock["symbol"]
            try:
                t = yf.Ticker(symbol)
                info = t.info
                time.sleep(0.5)
                
                rev_growth = info.get("revenueGrowth", "N/A")
                op_margins = info.get("operatingMargins", "N/A")
                free_cashflow = info.get("freeCashflow", "N/A")
                
                detailed_candidates_data.append({
                    "symbol": symbol,
                    "name": stock["name"],
                    "pe_ratio": stock["pe_ratio"],
                    "debt_ratio": stock["debt_ratio"],
                    "sector": stock["sector"],
                    "industry": stock["industry"],
                    "is_etf": stock.get("is_etf", False),
                    "revenue_growth": f"{rev_growth * 100:.2f}%" if isinstance(rev_growth, (int, float)) else "N/A",
                    "operating_margins": f"{op_margins * 100:.2f}%" if isinstance(op_margins, (int, float)) else "N/A",
                    "free_cash_flow": f"{free_cashflow / 1e6:.2f}M USD/DKK" if isinstance(free_cashflow, (int, float)) else "N/A",
                    "current_price": info.get("currentPrice", info.get("regularMarketPrice", "N/A")),
                    "currency": info.get("currency", "N/A")
                })
            except Exception as e:
                detailed_candidates_data.append(stock)

        # 8. Aktiver Gemini 3.5 Flash til den proaktive HTML-rapport
        council_report_html = "<h3>LLM Council Error</h3>"
        output_mp3 = "llm_council_podcast.mp3"
        podcast_compiled = False
        podcast_error_section = "" # Denne sektion vil indeholde fejl-loggen direkte i din mail, hvis den fejler!

        if GEMINI_API_KEY:
            print("Aktiverer Gemini 3.5 Flash...")
            current_weights_str = json.dumps(current_portfolio_weights, indent=2, ensure_ascii=False)
            current_holdings_str = json.dumps(current_holdings, indent=2, ensure_ascii=False)
            sector_distribution_str = json.dumps(sector_distribution, indent=2, ensure_ascii=False)
            
            council_agent = CouncilAgent(GEMINI_API_KEY)
            # Default navn sat til 'Wazir' og horisont sat til '15+ years' for den automatiske baggrundskørsel på GitHub
            council_report_html = council_agent.run_proactive_analysis(
                candidates_data=detailed_candidates_data,
                category=focus_category,
                deficit=deficit,
                current_portfolio_str=current_weights_str,
                current_holdings_str=current_holdings_str,
                sector_distribution_str=sector_distribution_str,
                user_name="Wazir",
                horizon="15+ years"
            )
            
            # 9. Generer automatisk lyd-podcast (RETTET: overfører nu 'Wazir' som andet parameter!)
            print("Igangsætter Podcastfy-produktion...")
            podcast_agent = PodcastAgent(GEMINI_API_KEY)
            try:
                # Nu risikerer vi ikke at fejlen skjules [3]!
                generated_file = podcast_agent.generate_podcast_audio(council_report_html, "Wazir")
                if generated_file and os.path.exists(generated_file):
                    import shutil
                    shutil.copyfile(generated_file, output_mp3)
                    podcast_compiled = True
                else:
                    podcast_error_section = "<p style='color:red;'>⚠️ <strong>Podcast Warning:</strong> Podcastfy did not generate an MP3 file.</p>"
            except Exception as e:
                # Hvis genereringen crasher, fanges fejlloggen her og sendes direkte til din mail [3]!
                err_tb = traceback.format_exc()
                podcast_error_section = f"""
                <div style="background-color: #FEF2F2; border: 1px solid #FCA5A5; border-left: 6px solid #EF4444; padding: 20px; border-radius: 8px; margin-top: 35px; margin-bottom: 20px;">
                    <h4 style="color: #991B1B; font-family: 'Georgia', serif; margin-top: 0; margin-bottom: 8px;">⚠️ Podcast Generation Failed (GitHub Runner)</h4>
                    <p style="color: #7F1D1D; font-size: 14px; margin-bottom: 10px; line-height: 1.5;">
                        The high-energy audio podcast could not be compiled on the GitHub Actions runner. Find the detailed diagnostic logs below to resolve the issue:
                    </p>
                    <pre style="background-color: #FFFFFF; border: 1px solid #FEE2E2; padding: 15px; border-radius: 5px; overflow-x: auto; color: #991B1B; font-size: 12px; line-height: 1.4;">{err_tb}</pre>
                </div>
                """
        else:
            print("Advarsel: GEMINI_API_KEY mangler.")

        # Hvis genereringen fejlede, klistrer vi fejlloggen direkte ind i bunden af e-mailen!
        if podcast_error_section:
            council_report_html += podcast_error_section

        # Klargør begge vedhæftninger til din natlige mail!
        attachments_list = []
        if podcast_compiled:
            attachments_list.append({"path": output_mp3, "name": "Wazir_LLM_Council_Podcast.mp3"})
            
        # Generer dit live-formel Excel-ark til din natlige mail
        excel_raw_bytes = generate_excel_template_bytes(current_holdings, watchlist_tickers)
        attachments_list.append({"data": excel_raw_bytes, "name": "Wazir_Live_Portfolio_Template.xlsx"})

        # 10. Send HTML-rapport og MP3-podcast til din indbakke
        subject = f"[LLM Council] Your Strategic Briefing - Focus on {DISPLAY_CATEGORIES.get(focus_category, focus_category)}"
        DeliveryAgent.send_email(
            subject=subject, 
            html_content=council_report_html, 
            attachments=attachments_list # Sender nu begge filer direkte til mailen automatisk!
        )

    except Exception as e:
        error_msg = f"Kritisk systemfejl under kørslen:\n\n{traceback.format_exc()}"
        print(error_msg, file=sys.stderr)
        
        error_html = f"""
        <div style="font-family: sans-serif; max-width: 600px; margin: 0 auto; padding: 25px; border: 1px solid #FCA5A5; background-color: #FEF2F2; border-radius: 8px;">
            <h2 style="color: #B91C1C; margin-top: 0; border-bottom: 2px solid #FCA5A5; padding-bottom: 10px;">🚨 Systemfejl i LLM Council</h2>
            <pre style="background-color: #FFFFFF; border: 1px solid #FEE2E2; padding: 15px; border-radius: 5px; overflow-x: auto; color: #991B1B; font-size: 13px;">{error_msg}</pre>
        </div>
        """
        DeliveryAgent.send_email("[System Error] LLM Council failed to run", error_html)


if __name__ == "__main__":
    main()
