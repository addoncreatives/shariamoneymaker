import os
import sys
import time
import io
import re
import json
import datetime
import traceback
import smtplib
import asyncio
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# ---------------------------------------------------------------------
#  SAFETY NET: Automatic installation of openpyxl if missing
# ---------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    import subprocess
    print("Safety net: openpyxl is missing. Installing automatically...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ---------------------------------------------------------------------
#  SAFETY NET: Automatic installation of edge-tts if missing
# ---------------------------------------------------------------------
try:
    import edge_tts
except ImportError:
    import subprocess
    print("Safety net: edge-tts is missing. Installing automatically...")
    subprocess.call([sys.executable, "-m", "pip", "install", "edge-tts"])
    import edge_tts

import yfinance as yf
import requests
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

# Page config
st.set_page_config(page_title="LLM Council - Conscious Wealth", page_icon="🗳️", layout="centered")


# =====================================================================
#  HTML & BILINGUAL TRANSLATION HELPERS
# =====================================================================
def render_html(html_str: str):
    clean_html = "".join([line.strip() for line in html_str.splitlines()])
    st.markdown(clean_html, unsafe_allow_html=True)


# Language selector (Danish, Swedish, Norwegian, Finnish, English)
col_title, col_lang = st.columns([2, 1])
with col_title:
    st.title("🗳️ LLM Council")
with col_lang:
    if "lang" not in st.session_state:
        st.session_state.lang = "Dansk"
    st.session_state.lang = st.selectbox(
        "Sprog / Language", 
        ["Dansk", "Svenska", "Norsk", "Suomi", "English"], 
        index=["Dansk", "Svenska", "Norsk", "Suomi", "English"].index(st.session_state.lang),
        label_visibility="collapsed"
    )

def _t(da: str, sv: str, no: str, fi: str, en: str) -> str:
    lang = st.session_state.lang
    if lang == "Dansk": return da
    elif lang == "Svenska": return sv
    elif lang == "Norsk": return no
    elif lang == "Suomi": return fi
    return en


# =====================================================================
#  CONFIGURATION & STANDARD TARGET WEIGHTS
# =====================================================================

DISPLAY_CATEGORIES = {
    "Aktier": "Equities",
    "Sukuk": "Sukuk (Islamic Bonds)",
    "Råvarer": "Commodities",
    "Kontanter/Private": "Cash / Private Sector"
}

UI_TO_DB_MAP = {
    "Equities": "Aktier",
    "Sukuk": "Sukuk",
    "Commodities": "Råvarer",
    "Cash/Private": "Kontanter/Private"
}

DB_TO_UI_MAP = {
    "Aktier": "Equities",
    "Sukuk": "Sukuk",
    "Råvarer": "Commodities",
    "Kontanter/Private": "Cash/Private"
}

TARGET_SUBSECTORS = [
    "Pharmaceuticals & Biotech",
    "Medical Devices & MedTech",
    "Clean Energy & Wind",
    "Smart Grid & Electrification",
    "Global Infrastructure",
    "Construction & Building Materials",
    "Chemicals & Advanced Materials",
    "Industrial Machinery & Automation",
    "Semiconductors & Hardware",
    "Mining & Royalty Streams",
    "Traditional Energy & Utilities",
    "Global Equity ETFs",
    "Regional & Thematic ETFs",
    "Sukuk & Fixed Income",
    "Cash & Liquidity Reserves",
    "Private Equity & Venture Capital",
    "Agriculture & Food Security",
    "Consumer Defensive & Staples",
    "Enterprise Software & SaaS",
    "Industrial Metals & Copper",
    "Logistics & Global Shipping",
    "Artificial Intelligence & Cloud Computing",
    "Cybersecurity & Digital Defense",
    "E-Commerce & Digital Retail",
    "Water Infrastructure & Desalination"
]

# Det statiske, lynhurtige kartotek (KID-fri og klar til Saxo)
STATIC_TICKER_MAP = {
    "NOVO-B.CO": ("Aktier", "Pharmaceuticals & Biotech"),
    "NOVO-B": ("Aktier", "Pharmaceuticals & Biotech"),
    "MSFT": ("Aktier", "Enterprise Software & SaaS"),
    "AAPL": ("Aktier", "Semiconductors & Hardware"),
    "SAP": ("Aktier", "Enterprise Software & SaaS"),
    "IFX.DE": ("Aktier", "Semiconductors & Hardware"),
    "ASML": ("Aktier", "Semiconductors & Hardware"),
    "NVDA": ("Aktier", "Semiconductors & Hardware"),
    "VWS.CO": ("Aktier", "Clean Energy & Wind"),
    "NKT.CO": ("Aktier", "Smart Grid & Electrification"),
    "FLS.CO": ("Aktier", "Industrial Machinery & Automation"),
    "ROCK-B.CO": ("Aktier", "Construction & Building Materials"),
    "ORK.OL": ("Aktier", "Consumer Defensive & Staples"),
    "WPM": ("Råvarer", "Mining & Royalty Streams"),
    "NEM": ("Råvarer", "Industrial Metals & Copper"),
    "AEM": ("Råvarer", "Industrial Metals & Copper"),
    "RGLD": ("Råvarer", "Mining & Royalty Streams"),
    "FSUK.L": ("Sukuk", "Sukuk & Fixed Income"),
    "SGLN.L": ("Råvarer", "Mining & Royalty Streams"),
    "MSAU.L": ("Aktier", "Regional & Thematic ETFs"),
    "IGDA.L": ("Aktier", "Global Equity ETFs"),
    "ISWD.DE": ("Aktier", "Global Equity ETFs"),
    "ISUS.DE": ("Aktier", "Regional & Thematic ETFs"),
    "HIWS.L": ("Aktier", "Global Equity ETFs")
}

@st.cache_data
def load_global_db_from_github():
    url = "https://raw.githubusercontent.com/addoncreatives/shariamoneymaker/main/failsafe_db.json"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            loaded_db = response.json()
            merged_db = STATIC_TICKER_MAP.copy()
            merged_db.update(loaded_db)
            return merged_db
    except Exception:
        pass
    return STATIC_TICKER_MAP

failsafe_db = load_global_db_from_github()

GLOBAL_COMPLIANT_GROWTH_POOL = {
    "Aktier": [
        "IGDA.L", "ISWD.DE", "ISUS.DE", "MSAU.L", "HIWS.L",
        "TRMB", "SAP", "IFX.DE", "MSFT", "ASML", "NVDA", "ADBE", "CRM", "SNPS", 
        "NOVO-B.CO", "6869.T", "AZN.ST", "REGN", "ISRG", "LLY", "VRTX", 
        "VWS.CO", "NKT.CO", "FLS.CO", "ROCK-B.CO"
    ],
    "Sukuk": [
        "FSUK.L"
    ],
    "Råvarer": [
        "SGLN.L", "WPM", "NEM", "GOLD", "AEM", "FNV", "RGLD", "BHP", "RIO", "FCX", "VALE"
    ],
    "Kontanter/Private": [
        "FSUK.L"
    ]
}

STATIC_ETF_CLONES = {
    "HLAL": [
        {"Company Name": "Microsoft Corp", "Ticker": "MSFT", "Shares": 5, "Category": "Aktier", "Sector": "Enterprise Software & SaaS", "Kurs": 0.0},
        {"Company Name": "Apple Inc", "Ticker": "AAPL", "Shares": 5, "Category": "Aktier", "Sector": "Semiconductors & Hardware", "Kurs": 0.0},
        {"Company Name": "NVIDIA Corp", "Ticker": "NVDA", "Shares": 4, "Category": "Aktier", "Sector": "Semiconductors & Hardware", "Kurs": 0.0},
        {"Company Name": "Alphabet Inc (Google)", "Ticker": "GOOGL", "Shares": 3, "Category": "Aktier", "Sector": "Artificial Intelligence & Cloud Computing", "Kurs": 0.0},
        {"Company Name": "Salesforce Inc", "Ticker": "CRM", "Shares": 2, "Category": "Aktier", "Sector": "Enterprise Software & SaaS", "Kurs": 0.0}
    ]
}

GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "1EnE2XkQySaGsdaxR5KySZZ924LT66ICo")
DATABASE_URL = os.getenv("DATABASE_URL")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

EMAIL_SENDER = os.getenv("EMAIL_SENDER", "wazir.ilyas@gmail.com")
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER", "addoncreatives@gmail.com")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
if EMAIL_PASSWORD:
    EMAIL_PASSWORD = EMAIL_PASSWORD.replace(" ", "").strip()


def normalize_string(s: str) -> str:
    if not s or pd.isna(s):
        return ""
    s = str(s).lower().strip()
    s = s.replace("og", "and").replace("&", "and")
    s = s.replace("'", "")
    s = re.sub(r'[^a-z0-9æøå]', '', s)
    s = s.replace("etfer", "etf").replace("etfs", "etf")
    return s


def search_tickers_by_name_multi(query: str) -> list:
    if not query or pd.isna(query) or len(str(query).strip()) < 2:
        return []
    
    query_clean = str(query).strip()
    url = f"https://query2.finance.yahoo.com/v1/finance/search?q={query_clean}"
    headers = {'User-Agent': 'Mozilla/5.0'}
    try:
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code == 200:
            data = response.json()
            quotes = data.get("quotes", [])
            results = []
            for q in quotes:
                sym = q.get("symbol")
                name = q.get("longname") or q.get("shortname") or sym
                q_type = q.get("quoteType", "").upper()
                if sym and q_type in ["EQUITY", "ETF", "MUTUALFUND"]:
                    results.append({"symbol": sym, "name": name})
            return results[:5]
    except Exception:
        pass
    return []


# =====================================================================
#  PASSIVE CASH-FLOW REBALANCING CALCULATOR
# =====================================================================
def calculate_passive_rebalancing_distribution(holdings_list: list, target_allocations: dict, monthly_deposit: float) -> dict:
    distribution_results = {"Aktier": 0.0, "Sukuk": 0.0, "Råvarer": 0.0, "Kontanter/Private": 0.0}
    if monthly_deposit <= 0:
        return distribution_results
        
    total_mv = 0.0
    category_values = {"Aktier": 0.0, "Sukuk": 0.0, "Råvarer": 0.0, "Kontanter/Private": 0.0}
    
    for item in holdings_list:
        category = item.get("Category", "Aktier")
        symbol = item.get("Ticker", "Other")
        if "PVT_" in symbol or "CASH_" in symbol:
            item_mv = float(item.get("manual_value", 1000))
        else:
            kurs = item.get("Kurs", 0.0)
            if kurs <= 0.0:
                try:
                    t = yf.Ticker(symbol)
                    kurs = t.info.get("currentPrice", t.info.get("regularMarketPrice", 1.0))
                except Exception:
                    kurs = 1.0
            item_mv = (kurs * float(item.get("Shares", 1)))
            
        total_mv += item_mv
        if category in category_values:
            category_values[category] += item_mv

    new_total_mv = total_mv + monthly_deposit
    
    deficits = {}
    total_deficit = 0.0
    for cat, target_pct in target_allocations.items():
        new_target_val = new_total_mv * (target_pct / 100.0)
        current_val = category_values.get(cat, 0.0)
        diff = new_target_val - current_val
        if diff > 0:
            deficits[cat] = diff
            total_deficit += diff
        else:
            deficits[cat] = 0.0

    if total_deficit > 0:
        for cat, diff in deficits.items():
            distribution_results[cat] = (diff / total_deficit) * monthly_deposit
    else:
        for cat, target_pct in target_allocations.items():
            distribution_results[cat] = (target_pct / 100.0) * monthly_deposit
            
    return distribution_results


# =====================================================================
#  HTML DEBATE PARSER (OVERSAT TIL SWOT HOVEDKASSER)
# =====================================================================
def parse_debate_and_chairman(html_content: str) -> dict:
    results = {
        "Contrarian": "",
        "First-Principles": "",
        "Expansionist": "",
        "Outsider": "",
        "Executor": "",
        "Chairman": ""
    }
    if not html_content:
        return results
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html_content, "html.parser")
        divs = soup.find_all("div")
        for d in divs:
            style = d.get("style", "")
            text = d.get_text().strip()
            if not text:
                continue
            if "EF4444" in style or "EF4444" in text or "Contrarian" in text[:45]:
                results["Contrarian"] = text
            elif "64748B" in style or "64748B" in text or "First-Principles" in text[:45] or "First Principles" in text[:45]:
                results["First-Principles"] = text
            elif "10B981" in style or "10B981" in text or "Expansionist" in text[:45]:
                results["Expansionist"] = text
            elif "8B5CF6" in style or "8B5CF6" in text or "Outsider" in text[:45]:
                results["Outsider"] = text
            elif "3B82F6" in style or "3B82F6" in text or "Executor" in text[:45]:
                results["Executor"] = text
            elif "C5A880" in style or "C5A880" in text or "Chairman" in text[:45] or "Marcus" in text[:45]:
                results["Chairman"] = text
    except Exception as e:
        print(f"BeautifulSoup parsing failed: {str(e)}")
    return results


# =====================================================================
#  EXCEL GENERATOR
# =====================================================================
def generate_excel_template_bytes(holdings_list: list, watchlist_list: list, portfolio_weights: dict = None, sector_distribution: dict = None) -> bytes:
    wb = openpyxl.Workbook()
    
    # 1. Beholdninger
    ws1 = wb.active
    ws1.title = "Beholdninger"
    
    headers1 = [
        "Position", "Ticker", "Status", "Antal", "Kurs (DKK)", 
        "Markedsværdi (DKK)", "Aktivklasse", "Drivkraft", "Sektor", 
        "Region", "Porteføljevægt", "Rolle", "Kommentar / tese"
    ]
    ws1.append(headers1)
    
    for idx, item in enumerate(holdings_list, start=2):
        name = item.get("Company Name", "Other")
        symbol = item.get("Ticker", "Other")
        shares = item.get("Shares", 1)
        kurs = item.get("Kurs", 0.0)
        cat = item.get("Category", "Aktier")
        sec = item.get("Sector", "Other")
        
        if "PVT_" in symbol or "CASH_" in symbol:
            val = float(item.get("manual_value", 1000))
            ws1.cell(row=idx, column=1, value=name)
            ws1.cell(row=idx, column=2, value="")
            ws1.cell(row=idx, column=3, value="Ejer")
            ws1.cell(row=idx, column=4, value=1)
            ws1.cell(row=idx, column=5, value=val)
            ws1.cell(row=idx, column=6, value=val)
        else:
            ws1.cell(row=idx, column=1, value=name)
            ws1.cell(row=idx, column=2, value=symbol)
            ws1.cell(row=idx, column=3, value="Ejer")
            ws1.cell(row=idx, column=4, value=shares)
            
            if kurs > 0.0:
                ws1.cell(row=idx, column=5, value=kurs)
            else:
                ws1.cell(row=idx, column=5, value=f'=GOOGLEFINANCE(B{idx})')
                
            ws1.cell(row=idx, column=6, value=f'=D{idx}*E{idx}')
            
        ws1.cell(row=idx, column=7, value=cat)
        ws1.cell(row=idx, column=8, value="")
        ws1.cell(row=idx, column=9, value=sec)
        ws1.cell(row=idx, column=10, value="Global")
        ws1.cell(row=idx, column=11, value=f'=F{idx}/SUM(F$2:F$100)')
        ws1.cell(row=idx, column=12, value="")
        ws1.cell(row=idx, column=13, value="")

    # 2. Opsummering
    ws2 = wb.create_sheet(title="Opsummering")
    ws2.cell(row=1, column=1, value="4x25-overblik")
    ws2.cell(row=1, column=6, value="Økonomiske drivere")
    ws2.cell(row=1, column=10, value="Sektorer")
    ws2.cell(row=1, column=14, value="Huller / Watchlist")
    
    ws2.cell(row=2, column=1, value="Kategori")
    ws2.cell(row=2, column=2, value="Aktuel Vægt (%)")
    ws2.cell(row=2, column=3, value="Mål Vægt (%)")
    
    ws2.cell(row=2, column=10, value="Delsektor")
    ws2.cell(row=2, column=11, value="Vægt (%)")
    ws2.cell(row=2, column=14, value="Ticker")

    if portfolio_weights:
        categories = ["Aktier", "Sukuk", "Råvarer", "Kontanter/Private"]
        for i_idx, cat in enumerate(categories, start=3):
            ws2.cell(row=i_idx, column=1, value=cat)
            ws2.cell(row=i_idx, column=2, value=portfolio_weights.get(cat, 0.0))
            ws2.cell(row=i_idx, column=3, value=25.0)

    if sector_distribution:
        for s_idx, (sec, weight) in enumerate(sector_distribution.items(), start=3):
            ws2.cell(row=s_idx, column=10, value=sec)
            ws2.cell(row=s_idx, column=11, value=weight)

    for w_idx, ticker in enumerate(watchlist_list, start=3):
        ws2.cell(row=w_idx, column=14, value=ticker)
        
    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data.getvalue()


def get_category_and_sector_failsafe(ticker: str, target_category: str = None) -> tuple:
    sym = str(ticker).upper().strip()
    lookup_sym = sym.split('.')[0]
    
    for k, v in failsafe_db.items():
        if normalize_string(k) == normalize_string(sym) or normalize_string(k) == normalize_string(lookup_sym):
            if v[0] == "Sukuk" and target_category == "Kontanter/Private":
                return "Kontanter/Private", "Sukuk & Fixed Income"
            return v[0], v[1]
            
    try:
        t = yf.Ticker(sym)
        info = t.info
        sec = info.get("sector", "Other")
        ind = info.get("industry", "Other")
        
        temp_screener = ScreenerComplianceAgent([], target_category=target_category)
        cat, sub_sec = temp_screener.map_to_category_and_sector(sym, sec, ind)
        return cat, sub_sec
    except Exception:
        return "Aktier", "Other"


# =====================================================================
#  SÆRLIG SCREENING & AGENTER
# =====================================================================
class PortfolioManagerAgent:
    def __init__(self, current: dict, target: dict):
        self.current = current
        self.target = target

    def identify_underweighted_focus(self) -> tuple:
        underweight_candidates = []
        for category, target_val in self.target.items():
            curr_val = self.current.get(category, 0.0)
            deficit = target_val - curr_val
            if deficit > 0.0:
                underweight_candidates.append((category, deficit))
        
        if not underweight_candidates:
            return list(self.target.keys())[0], 0.0

        underweight_candidates.sort(key=lambda x: x[0])
        day_of_year = datetime.datetime.now().timetuple().tm_yday
        index = day_of_year % len(underweight_candidates)
        
        focus_category, deficit = underweight_candidates[index]
        return focus_category, deficit


class ScreenerComplianceAgent:
    PROHIBITED_SECTORS = ["Financial Services", "Financial"]
    PROHIBITED_INDUSTRIES = ["Banks", "Insurance", "Aerospace & Defense", "Gambling", "Tobacco", "Distillers & Vintners", "Breweries"]

    def __init__(self, tickers: list, target_category: str = None):
        self.tickers = tickers
        self.target_category = target_category

    def check_zoya_live_compliance(self, symbol: str) -> bool:
        clean_symbol = symbol.split('.')[0].upper().split('-')[0]
        url = f"https://zoya.finance/stocks/{clean_symbol}"
        headers = {'User-Agent': 'Mozilla/5.0'}
        try:
            response = requests.get(url, headers=headers, timeout=8)
            if response.status_code == 200:
                html_lower = response.text.lower()
                if "is shariah-compliant" in html_lower or "is considered halal" in html_lower:
                    return True
                elif "not shariah-compliant" in html_lower or "non-compliant" in html_lower:
                    return False
            return None
        except Exception:
            return None

    def map_to_category_and_sector(self, symbol: str, sector: str = None, industry: str = None) -> tuple:
        sym = symbol.upper().strip()
        sec_l = sector.lower() if sector else ""
        ind_l = industry.lower() if industry else ""

        lookup_sym = sym.split('.')[0]
        for k, v in failsafe_db.items():
            if normalize_string(k) == normalize_string(sym) or normalize_string(k) == normalize_string(lookup_sym):
                if v[0] == "Sukuk" and self.target_category == "Kontanter/Private":
                    return "Kontanter/Private", "Sukuk & Fixed Income"
                return v[0], v[1]

        if "sukuk" in sym or sym in ["SPSK", "SKUK", "FSUK.L"]:
            if self.target_category == "Kontanter/Private":
                return "Kontanter/Private", "Sukuk & Fixed Income"
            return "Sukuk", "Sukuk & Fixed Income"
            
        if sym in ["WPM", "FNV", "RGLD"]:
            return "Råvarer", "Mining & Royalty Streams"
        if sym in ["NEM", "GOLD", "AEM", "BHP", "RIO", "FCX", "VALE", "SGLN.L"] or any(w in ind_l for w in ["gold", "silver", "precious metals", "copper", "aluminum"]):
            return "Råvarer", "Industrial Metals & Copper"

        if "cash" in sym or "money market" in sec_l:
            return "Kontanter/Private", "Cash & Liquidity Reserves"

        if "pharmaceutical" in ind_l or "biotechnology" in ind_l:
            return "Aktier", "Pharmaceuticals & Biotech"
        if "medical" in ind_l or "healthcare" in sec_l:
            return "Aktier", "Medical Devices & MedTech"
        if "wind" in ind_l or "renewable" in ind_l or "solar" in ind_l:
            return "Aktier", "Clean Energy & Wind"
        if "cable" in ind_l or "electrical" in ind_l:
            return "Aktier", "Smart Grid & Electrification"
        if "semiconductor" in ind_l or "semiconductor" in sec_l:
            return "Aktier", "Semiconductors & Hardware"
        if "software" in ind_l or "software" in sec_l or "technology" in sec_l:
            return "Aktier", "Enterprise Software & SaaS"
        if "building" in ind_l or "construction" in ind_l:
            return "Aktier", "Construction & Building Materials"
        if "chemicals" in ind_l:
            return "Aktier", "Chemicals & Advanced Materials"
        if "machinery" in ind_l or "industrials" in sec_l:
            return "Aktier", "Industrial Machinery & Automation"
        if "infrastructure" in ind_l or "utilities" in sec_l:
            return "Aktier", "Global Infrastructure"
        if "staples" in sec_l or "packaged foods" in ind_l or "consumer defensive" in sec_l:
            return "Aktier", "Consumer Defensive & Staples"
        if "fertilizer" in ind_l or "agriculture" in ind_l:
            return "Aktier", "Agriculture & Food Security"
        if "logistics" in ind_l or "shipping" in ind_l:
            return "Aktier", "Logistics & Global Shipping"
        if "internet" in ind_l or "e-commerce" in ind_l:
            return "Aktier", "E-Commerce & Digital Retail"
        if "water" in ind_l or "environmental" in ind_l:
            return "Aktier", "Water Infrastructure & Desalination"

        dynamic_subsector = industry if (industry and industry != "Other") else sector
        return "Aktier", dynamic_subsector

    def screen_ticker(self, symbol: str) -> dict:
        try:
            if "CASH_" in symbol or "PVT_" in symbol:
                return {
                    "symbol": symbol, "passed": True, "name": symbol.replace("CASH_", "").replace("PVT_", ""),
                    "pe_ratio": "N/A", "debt_ratio": "0.00% (Manual)", "sector": "Manual Asset",
                    "industry": "Manual Asset", "category": "Kontanter/Private", "subsector": "Cash & Liquidity Reserves", "is_etf": False
                }

            zoya_compliant = self.check_zoya_live_compliance(symbol)
            if zoya_compliant is False:
                return {"symbol": symbol, "passed": False, "reason": "Disqualified by Zoya assessment."}

            ticker_obj = yf.Ticker(symbol)
            info = ticker_obj.info
            if not info:
                return {"symbol": symbol, "passed": False, "reason": "Ingen data"}

            quote_type = info.get("quoteType", "").upper()
            is_etf = quote_type in ["ETF", "MUTUALFUND"] or symbol in ["IGDA.L", "SPSK", "HLAL", "UMMA", "ISWD.L", "MSAU.L", "SKUK", "FSUK.L", "ISWD.DE", "ISUS.DE", "SGLN.L"]

            if is_etf:
                mapped_cat, mapped_sub = self.map_to_category_and_sector(symbol, "ETF", "ETF")
                return {
                    "symbol": symbol, "passed": True, "name": info.get("longName", symbol),
                    "pe_ratio": info.get("trailingPE", "N/A"), "debt_ratio": "N/A (ETF/Sukuk)",
                    "sector": "ETF / Fond", "industry": "ETF", "category": mapped_cat, "subsector": mapped_sub, "is_etf": True
                }

            sector = info.get("sector", "")
            industry = info.get("industry", "")
            
            for p_sector in self.PROHIBITED_SECTORS:
                if p_sector.lower() in sector.lower():
                    return {"symbol": symbol, "passed": False, "reason": f"Sektor: {sector}"}
            for p_ind in self.PROHIBITED_INDUSTRIES:
                if p_ind.lower() in industry.lower():
                    return {"symbol": symbol, "passed": False, "reason": f"Branche: {industry}"}

            debt_to_equity = info.get("debtToEquity")
            total_debt = info.get("totalDebt")
            market_cap = info.get("marketCap")
            
            debt_ratio_pct = None
            method_used = ""
            if debt_to_equity is not None:
                debt_ratio_pct = debt_to_equity
                method_used = "Debt to Equity"
            elif total_debt and market_cap:
                debt_ratio_pct = (total_debt / market_cap) * 100
                method_used = "Debt to Market Cap"

            if debt_ratio_pct is None:
                return {"symbol": symbol, "passed": False, "reason": "Ingen gældsdata"}
            if debt_ratio_pct > 30.0:
                return {"symbol": symbol, "passed": False, "reason": f"Gældskvote: {debt_ratio_pct:.2f}%"}

            mapped_cat, mapped_sub = self.map_to_category_and_sector(symbol, sector, industry)
            return {
                "symbol": symbol, "passed": True, "name": info.get("longName", symbol),
                "pe_ratio": info.get("trailingPE", "N/A"), "debt_ratio": f"{debt_ratio_pct:.2f}% ({method_used})",
                "sector": sector, "industry": industry, "category": mapped_cat, "subsector": mapped_sub, "is_etf": False
            }
        except Exception as e:
            return {"symbol": symbol, "passed": False, "reason": str(e)}

    def run_screening(self, target_category: str) -> list:
        approved_stocks = []
        for ticker in self.tickers:
            time.sleep(1.5)
            result = self.screen_ticker(symbol=ticker)
            if result["passed"] and result["category"] == target_category:
                approved_stocks.append(result)
        return approved_stocks


# =====================================================================
#  COUNCIL AGENT
# =====================================================================
class CouncilAgent:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-3.5-flash:generateContent?key={self.api_key}"

    def run_proactive_analysis(self, candidates_data: list, category: str, deficit: float, current_portfolio_str: str, current_holdings_str: str, sector_distribution_str: str, user_name: str, horizon: str) -> str:
        candidates_json = json.dumps(candidates_data, indent=2, ensure_ascii=False)
        english_category = DISPLAY_CATEGORIES.get(category, category)
        
        prompt = f"""
        You are an elite financial advisory council ("LLM Council") presenting a strategic investment briefing to your highly valued VIP client, {user_name}.
        
        THE INVESTOR PROFILE & MODEL:
        - Investor's Name: {user_name}
        - Investment Horizon: {horizon}
        - Focus tonight: {english_category} (Current Deficit: {deficit:.2f}%)
        - Current Portfolio Allocations: {current_portfolio_str}
        - Sub-sectors: {sector_distribution_str}
        - Current Holdings: {current_holdings_str}
        - Screened candidates: {candidates_json}
        
        YOUR OBJECTIVE:
        Generate a complete, institutional-grade, highly engaging HTML investment newsletter in English with inline CSS styling.
        Refer to AAOIFI ethical standards (under 30% debt rule). Avoid extreme financial jargon. Frame recommendations respectfully.
        """
        headers = {'Content-Type': 'application/json'}
        payload = {"contents": [{"parts": [{"text": prompt}]}]}
        try:
            response = requests.post(self.url, headers=headers, json=payload, timeout=110)
            response.raise_for_status()
            data = response.json()
            if 'candidates' in data and len(data['candidates']) > 0:
                raw_html = data['candidates'][0]['content']['parts'][0]['text']
                return raw_html.replace("```html", "").replace("```", "").strip()
        except Exception as e:
            return f"<h3>System Error</h3><p>{str(e)}</p>"


# =====================================================================
#  ON-DEMAND SEKTOR PROSPEKTOR AGENT (SWOT MODEL INDBYGGET)
# =====================================================================
def generate_sector_prospects(api_key: str, sector: str, user_name: str, ignore_list: list = None) -> str:
    """
    Genererer 3 Shariah-compliant selskaber inden for den valgte delsektor 
    via Gemini og returnerer dem som en rå, struktureret JSON-streng med meninger fra rådet.
    """
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    
    ignore_clause = ""
    if ignore_list:
        ignore_clause = f"CRITICAL: Exclude the following tickers from your analysis: {', '.join(ignore_list)}."

    prompt = f"""
    You are the elite financial advisory council ("LLM Council"). Our VIP client, {user_name}, has engaged you to research and prospect the sector: '{sector}'.
    {ignore_clause}
    
    YOUR TASK:
    1. Identify 3 promising, Shariah-compliant (based on standard AAOIFI interest-debt criteria) global companies in the '{sector}' sector.
    2. Format the response strictly as a raw JSON array of 3 objects. Do NOT use markdown code blocks like "```json". Just output the raw JSON.
    
    JSON Schema:
    [
      {{
        "symbol": "TICKER",
        "name": "Company Name",
        "thesis": "A short, highly engaging, storytelling-focused investment case (maximum 2 sentences) in English. Focus on future growth pipeline, not financial numbers.",
        "sa_link": "https://seekingalpha.com/symbol/TICKER",
        "yf_link": "https://finance.yahoo.com/quote/TICKER",
        "ir_link": "https://www.google.com/search?q=TICKER+investor+relations",
        "kid_status": "stock",
        "opinions": {{
          "contrarian": "A short (2-3 sentences) critical, highly skeptical, risk-obsessed argument warning of margins, valuations, or competitor moats.",
          "expansionist": "A short (2-3 sentences) highly bullish argument focusing on growth potential, secure cash flow, and market-leading momentum.",
          "outsider": "A short (2-3 sentences) macro argument linking the company to secular global megatrends and geopolitical dynamics.",
          "first_principles": "A short (2-3 sentences) analysis of Shariah debt-compliance under strict AAOIFI rules (must stay under 30% debt-to-market-cap ratio).",
          "executor": "A short (2-3 sentences) pragmatic assessment of buy-and-hold tradeability on Saxo and Nordnet.",
          "chairman": "A short (2-3 sentences) final committee recommendation on how to purchase (e.g. recommend dollar-cost averaging over 3 months)."
        }}
      }}
    ]
    """
    headers = {'Content-Type': 'application/json'}
    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=120)
        response.raise_for_status()
        data = response.json()
        if 'candidates' in data and len(data['candidates']) > 0:
            return data['candidates'][0]['content']['parts'][0]['text']
    except Exception as e:
        print(f"Fejl under Sektor-prospektering: {str(e)}")
    return "[]"

def extract_json_array(text: str) -> list:
    """Metode til fejlfrit at trække og parse et JSON-array fra Gemini output."""
    match = re.search(r'\[\s*\{.*\}\s*\]', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            pass
    try:
        return json.loads(text.strip().replace("```json", "").replace("```", "").strip())
    except Exception:
        return []


# =====================================================================
#  FEJLSIKRET DYNAMISK ETF PURIFIER (KORRIGERET TIL BENHÅRDT SIKKERHEDS-TJEK)
# =====================================================================
def purify_conventional_etf(api_key: str, etf_ticker: str, top_n: int) -> list:
    """
    Kalder Gemini til at identificere de top-N største selskaber i en konventionel ETF,
    analysere deres Shariah-compliance (hvor forretningsmodellen samt rentebærende gæld
    mod markedsværdien under 30% (AAOIFI-regler) er det primære tjek),
    og returnere en JSON-liste med resultaterne.
    """
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    prompt = f"""
    You are the elite financial advisory council ("LLM Council") specializing in AAOIFI Shariah screening.
    The user wants to purify the conventional ETF with ticker '{etf_ticker}' by auditing its top {top_n} largest holdings.
    
    YOUR TASK (SHARIAH DEBT-AUDIT IS THE HIGHEST PRIORITY):
    1. Identify the top {top_n} largest individual stock holdings of the conventional ETF '{etf_ticker}' (e.g., if SPY, identify its top holdings).
    2. Audit each company against strict Shariah compliance and AAOIFI standards:
       - Business compliance (strictly no conventional financial services, banks, insurance, pork, alcohol, weapons, gambling, etc.).
       - Debt compliance (strictly no companies where interest-bearing debt exceeds 30% of their total market capitalization). If debt is high, you MUST mark them as Non-Compliant.
    3. Classify each holding as either "Compliant" or "Non-Compliant".
    4. Provide a very brief, 1-sentence reason for your decision in English (explicitly mention if they were purged due to high interest-bearing debt, banking services, etc.).
    5. Format your response strictly as a raw JSON array of objects (do NOT use markdown "```json" blocks):
    
    JSON Schema:
    [
      {{
        "symbol": "TICKER",
        "name": "Company Name",
        "status": "Compliant" or "Non-Compliant",
        "reason": "Brief reason in English...",
        "sa_link": "https://seekingalpha.com/symbol/TICKER",
        "yf_link": "https://finance.yahoo.com/quote/TICKER",
        "ir_link": "https://www.google.com/search?q=TICKER+investor+relations"
      }}
    ]
    """
    headers = {'Content-Type': 'application/json'}
    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=120)
        response.raise_for_status()
        data = response.json()
        if 'candidates' in data and len(data['candidates']) > 0:
            return extract_json_array(data['candidates'][0]['content']['parts'][0]['text'])
    except Exception as e:
        print(f"ETF Purifier fejlede: {str(e)}")
    return []


# =====================================================================
#  PODCAST AGENT (KORRIGERET TIL MULTI-ADVISOR SYNERGY OG RETTEDE XML-TAGS)
# =====================================================================
class PodcastAgent:
    def __init__(self, api_key: str):
        self.api_key = api_key
        if self.api_key:
            os.environ["GEMINI_API_KEY"] = self.api_key
            os.environ["GOOGLE_API_KEY"] = self.api_key

    def generate_podcast_audio(self, report_html: str, user_name: str) -> str:
        from podcastfy.client import generate_podcast
        
        user_instructions_content = (
            f"Set up a high-energy educational narrative podcast. "
            f"Explain to the listener our VIP client {user_name}'s current portfolio status and macro investment options. "
            f"Sara (Person 1) acts as the primary host. She must present the global macro events (such as supply chains or central banks) "
            f"and then discuss 2-3 Shariah-compliant candidates. "
            f"To present a structured critique, Sara must summarize the views of three virtual advisors using their names and titles:\n"
            f"- David, Chief Strategist at a global short-fund (skeptical and cautious)\n"
            f"- Michael, CIO of a Silicon Valley growth fund (enthusiastic about pipelines and growth)\n"
            f"- Elena, Head of Macro Research at a Swiss investment bank (focused on Shariah compliance limits and asset balance)\n\n"
            f"Marcus (Person 2), the Investment Committee Chairman, must join Sara in the second half of the episode. "
            f"Marcus must act as the Executor, explaining the practical tradeability on Saxo Bank and recommending a calm, step-by-step strategy. "
            f"Both hosts must summarize the final Masterclass verdict, keep 'number noise' to an absolute minimum, and make it educational and story-driven."
        )

        custom_config = {
            "word_count": 1100,
            "conversation_style": ["educational", "professional", "highly engaging", "storytelling", "structured dialogue"],
            "roles_person1": "Sara, the sharp financial journalist and solo host",
            "roles_person2": "Marcus, the wise Investment Committee Chairman",
            "podcast_name": "The Investor's Journey",
            "podcast_tagline": "Your money, your journey, your Shariah-compliant future",
            "output_language": "English",
            "engagement_techniques": ["rhetorical questions", "analogies", "humor", "interjections"],
            "user_instructions": user_instructions_content
        }
        
        try:
            print("Genererer struktureret podcast (Sara & Marcus) på engelsk...")
            audio_path = generate_podcast(
                text=report_html,
                tts_model="edge",
                conversation_config=custom_config
            )
            return audio_path
        except Exception as e:
            print(f"Fejl under Podcastfy-generering: {str(e)}")
            return None


# =====================================================================
#  DELIVERY AGENT
# =====================================================================
class DeliveryAgent:
    @staticmethod
    def send_email(subject: str, html_content: str, attachments: list = None):
        if not EMAIL_PASSWORD or EMAIL_PASSWORD.strip() == "":
            print(html_content)
            return

        msg = MIMEMultipart()
        msg["From"] = EMAIL_SENDER
        msg["To"] = EMAIL_RECEIVER
        msg["Subject"] = subject
        msg.attach(MIMEText(html_content, "html", "utf-8"))

        if attachments:
            for att in attachments:
                path = att.get("path")
                data = att.get("data")
                name = att.get("name")
                
                if (path and os.path.exists(path)) or data:
                    part = MIMEBase("application", "octet-stream")
                    if data:
                        part.set_payload(data)
                    else:
                        with open(path, "rb") as attachment:
                            part.set_payload(attachment.read())
                    
                    encoders.encode_base64(part)
                    part.add_header("Content-Disposition", f"attachment; filename= {name}")
                    msg.attach(part)

        try:
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.send_message(msg)
            server.quit()
        except Exception:
            pass


# =====================================================================
#  DATABASE INTEGRATION (GOOGLE WEB APP)
# =====================================================================
def load_user_portfolio_from_db(email: str, password: str) -> dict:
    if not DATABASE_URL or DATABASE_URL.strip() == "":
        return None
    try:
        response = requests.get(f"{DATABASE_URL}?email={email}&password={password}", timeout=10)
        if response.status_code == 200:
            data = response.json()
            if data.get("status") == "success":
                return data
    except Exception:
        pass
    return None

def save_user_portfolio_to_db(email: str, password: str, holdings: list, targets: dict, horizon: str, name: str, frequency: str) -> str:
    if not DATABASE_URL or DATABASE_URL.strip() == "":
        return "no_db"
    payload = {
        "email": email, "password": password, "holdings": holdings,
        "targets": targets, "horizon": horizon, "name": name, "frequency": frequency
    }
    try:
        response = requests.post(DATABASE_URL, json=payload, timeout=15)
        if response.status_code == 200:
            return response.json().get("status")
    except Exception:
        pass
    return "error"


# =====================================================================
#  SESSION STATE INITIALIZATION
# =====================================================================
if "step" not in st.session_state:
    st.session_state.step = 1
if "investor_holdings" not in st.session_state:
    st.session_state.investor_holdings = []
if "targets" not in st.session_state:
    st.session_state.targets = {"Aktier": 25.0, "Sukuk": 25.0, "Råvarer": 25.0, "Kontanter/Private": 25.0}
if "horizon" not in st.session_state:
    st.session_state.horizon = "7-15 years"
if "user_name" not in st.session_state:
    st.session_state.user_name = "Investor"
if "frequency" not in st.session_state:
    st.session_state.frequency = "Weekly"
if "user_email" not in st.session_state:
    st.session_state.user_email = ""
if "is_logged_in" not in st.session_state:
    st.session_state.is_logged_in = False
if "is_new_investor" not in st.session_state:
    st.session_state.is_new_investor = False
if "watchlist_list" not in st.session_state:
    st.session_state.watchlist_list = ["TRMB", "SAP", "SPSK", "AEM", "NEM"]

# Slidervariabler
if "slider_stocks" not in st.session_state:
    st.session_state.slider_stocks = int(st.session_state.targets.get("Aktier", 25.0))
if "slider_sukuk" not in st.session_state:
    st.session_state.slider_sukuk = int(st.session_state.targets.get("Sukuk", 25.0))
if "slider_commodities" not in st.session_state:
    st.session_state.slider_commodities = int(st.session_state.targets.get("Råvarer", 25.0))
if "slider_cash" not in st.session_state:
    st.session_state.slider_cash = int(st.session_state.targets.get("Kontanter/Private", 25.0))

# Failsafe lagring af resultater
if "generated_report" not in st.session_state:
    st.session_state.generated_report = None
if "generated_audio_bytes" not in st.session_state:
    st.session_state.generated_audio_bytes = None

# Sektor research
if "last_sector_prospects_list" not in st.session_state:
    st.session_state.last_sector_prospects_list = []
if "last_research_sector_name" not in st.session_state:
    st.session_state.last_research_sector_name = None
if "prospect_ignore_list" not in st.session_state:
    st.session_state.prospect_ignore_list = []

# Månedlig opsparing til passiv rebalancering
if "monthly_deposit" not in st.session_state:
    st.session_state.monthly_deposit = 2000.0

# ETF Purifier resultater
if "peeled_etf_list" not in st.session_state:
    st.session_state.peeled_etf_list = []
if "last_peeled_etf_ticker" not in st.session_state:
    st.session_state.last_peeled_etf_ticker = None


# =====================================================================
#  STREAMLIT UI STEPS
# =====================================================================

# Diskret status-indikator i toppen (Rettet sv/no sprogfejl!)
st.caption(_t(
    f"Trin {st.session_state.step} af 5", 
    f"Steg {st.session_state.step} av 5", 
    f"Trinn {st.session_state.step} av 5", 
    f"Vaihe {st.session_state.step} / 5", 
    f"Step {st.session_state.step} of 5"
))


# --- TRIN 1: VELKOMST & KORT KONTEKST (SAAS DASHBOARD FOR LOGGET-IND BRUGER) ---
if st.session_state.step == 1:
    if st.session_state.is_logged_in:
        # PERSONLIGT DASHBOARD FOR EKSISTERENDE BRUGERE
        st.subheader(_t(
            f"Velkommen tilbage, {st.session_state.user_name}!", 
            f"Välkommen tillbaka, {st.session_state.user_name}!", 
            f"Velkommen tilbake, {st.session_state.user_name}!", 
            f"Tervetuloa takaisin, {st.session_state.user_name}!", 
            f"Welcome back, {st.session_state.user_name}!"
        ))
        
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.metric(_t("Antal aktive positioner", "Antal aktiva positioner", "Antall aktive posisjoner", "Aktiiviset salkun osat", "Active Holdings"), len(st.session_state.investor_holdings))
        with col_d2:
            st.metric(_t("Briefing-frekvens", "Briefing-frekvens", "Briefing-frekvens", "Briefing-aikataulu", "Briefing Frequency"), st.session_state.frequency)

        # PASSIV CASH-FLOW REBALANCERING DIREKTE PÅ DASHBOARDET
        if st.session_state.investor_holdings and st.session_state.monthly_deposit > 0:
            st.write("---")
            st.subheader("💰 " + _t("Dit månedlige rebalancerings-budget", "Din månatliga ombalanseringsbudget", "Ditt månedlige rebalanceringsbudsjett", "Kuukausittainen tasapainotusbudjettisi", "Your Monthly Rebalancing Budget"))
            st.write(_t(
                f"Baseret på din planlagte månedlige opsparing på **{st.session_state.monthly_deposit:,.2f} DKK**, anbefaler rådet, at du fordeler dine *nye* indskud således for at lukke allokeringshullerne passivt helt uden salgskurtager:",
                f"Baserat på din månatliga sparning på **{st.session_state.monthly_deposit:,.2f} DKK** rekommenderar rådet att du fördelar dina nya insättningar så här för att stänga gappen passivt utan säljavgifter:",
                f"Basert på din månedlige sparing på **{st.session_state.monthly_deposit:,.2f} DKK** anbefaler rådet at du fordeler dine nye innskudd slik for å lukke hullene passivt uten salgskurtasjer:",
                f"Kuukausittaisen **{st.session_state.monthly_deposit:,.2f} DKK** säästösi perusteella neuvosto suosittelee, että jaat uudet talletuksesi seuraavasti tasapainottaaksesi salkkusi passiivisesti ilman myyntikuluja:",
                f"Based on your planned monthly deposit of **{st.session_state.monthly_deposit:,.2f} DKK**, the Council recommends distributing your new capital as follows to passively close the gaps without paying selling fees:"
            ))
            
            rebal_dist = calculate_passive_rebalancing_distribution(st.session_state.investor_holdings, st.session_state.targets, st.session_state.monthly_deposit)
            col_r1, col_r2, col_r3, col_r4 = st.columns(4)
            with col_r1:
                st.metric(_t("Køb Aktier", "Köp Aktier", "Kjøp Aksjer", "Osta Osakkeita", "Buy Equities"), f"{rebal_dist['Aktier']:,.0f} DKK")
            with col_r2:
                st.metric(_t("Køb Sukuk", "Köp Sukuk", "Kjøp Sukuk", "Osta Sukuk", "Buy Sukuk"), f"{rebal_dist['Sukuk']:,.0f} DKK")
            with col_r3:
                st.metric(_t("Køb Råvarer", "Köp Råvaror", "Kjøp Råvarer", "Osta Raaka-aineita", "Buy Commodities"), f"{rebal_dist['Råvarer']:,.0f} DKK")
            with col_r4:
                st.metric(_t("Gem i Kontanter", "Spara i Kontanter", "Spar i Kontanter", "Säästä Käteisenä", "Keep in Cash"), f"{rebal_dist['Kontanter/Private']:,.0f} DKK")

        # Hurtig oversigt over beholdninger
        if st.session_state.investor_holdings:
            st.write("---")
            st.subheader(_t("Din aktuelle portefølje", "Din nuvarande portfölj", "Din nåværende portefølje", "Nykyinen salkkusi", "Your Current Portfolio"))
            df_dash = pd.DataFrame(st.session_state.investor_holdings)
            df_dash['Category_Display'] = df_dash['Category'].apply(lambda x: DB_TO_UI_MAP.get(x, x))
            st.dataframe(
                df_dash,
                column_config={
                    "Company Name": _t("Selskabsnavn", "Bolagsnamn", "Selskapsnavn", "Yrityksen nimi", "Company Name"),
                    "Ticker": "Ticker",
                    "Shares": _t("Antal", "Antal", "Antall", "Määrä", "Shares"),
                    "Category_Display": _t("Kategori", "Kategori", "Kategori", "Kategoria", "Category"),
                    "Sector": _t("Delsektor", "Delsektor", "Delsektor", "Sektori", "Sub-sector")
                },
                use_container_width=True
            )

        # DET NYE DASHBOARD WATCHLIST MODUL (MED REELLE TICKERE)
        if st.session_state.watchlist_list:
            st.write("---")
            st.subheader("📊 " + _t("Din Watchlist", "Din Watchlist", "Din Watchlist", "Tarkkailulistasi", "Your Watchlist"))
            watchlist_data = []
            for ticker in st.session_state.watchlist_list:
                try:
                    t = yf.Ticker(ticker)
                    price = t.info.get("currentPrice", t.info.get("regularMarketPrice", "N/A"))
                    currency = t.info.get("currency", "")
                    watchlist_data.append({
                        "Ticker": ticker, 
                        "Price": f"{price} {currency}" if price != "N/A" else "N/A"
                    })
                except Exception:
                    watchlist_data.append({"Ticker": ticker, "Price": "N/A"})
            st.dataframe(pd.DataFrame(watchlist_data), use_container_width=True)
        
        st.write("---")
        # Genvejs-links til redigering eller rapport-kørsel
        st.subheader(_t("Hurtige handlinger", "Snabbval", "Hurtige handlinger", "Pika-asennot", "Quick Actions"))
        col_act1, col_act2 = st.columns(2)
        with col_act1:
            if st.button("⚙️ " + _t("Rediger portefølje & mål", "Redigera portfölj & mål", "Rediger portefølje & mål", "Muokkaa salkkua & tavoitteita", "Edit Portfolio & Targets"), use_container_width=True):
                st.session_state.step = 3
                st.rerun()
        with col_act2:
            if st.button("🗳️ " + _t("Generer ny LLM-rapport", "Generera ny LLM-rapport", "Generer ny LLM-rapport", "Luo uusi LLM-raportti", "Generate New Briefing"), use_container_width=True):
                st.session_state.step = 5
                st.rerun()

        # DYNAMISK ON-DEMAND SEKTOR RESEARCH (INTERAKTIVT DIVE-IN MED WATCHLIST TILFØJELSE)
        st.write("---")
        st.subheader("🔍 " + _t("Strategisk Rådgiver-Terminal", "Strategisk Rådgivarterminal", "Strategisk Rådgiverterminal", "Strateginen Neuvonantajapääte", "Strategic Advisor Terminal"))
        st.write(_t(
            "Vælg en delsektor nedenfor for at lade dine rådgivere foretage en øjeblikkelig prospektering af 3 stærke, compliant vækstcases. Du kan tilføje dem direkte til din Watchlist herunder.",
            "Välj en delsektor nedan för att låta dina rådgivare göra en omedelbar prospektering av 3 starka, compliant tillväxtcase. Du kan lägga till dem direkt i din Watchlist nedan.",
            "Velg en delsektor nedenfor for å la dine rådgivere foreta en dypgående prospektering av 3 sterke, compliant vekstcases. Du kan legge dem til i din Watchlist direkte nedenfor.",
            "Valitse alta osa-sektori, jotta neuvonantajat voivat tehdä välittömän arvion kolmesta vahvasta ja vaatimukset täyttävästä kasvukohteesta. Voit lisätä ne suoraan tarkkailulistallesi alta.",
            "Select a sub-sector below to have your advisors perform an on-demand prospecting of 3 strong, compliant growth cases. You can add them directly to your Watchlist below."
        ))
        
        col_sec_sel, col_sec_reload = st.columns([2, 1])
        with col_sec_sel:
            selected_research_sector = st.selectbox(
                _t("Vælg delsektor:", "Välj delsektor:", "Velg delsektor:", "Valitse osa-sektori:", "Select sub-sector to prospect:"),
                TARGET_SUBSECTORS
            )
        with col_sec_reload:
            st.write(" ")
            st.write(" ")
            # Knap til at hente alternative selskaber inden for samme sektor (Ignore-liste aktiveres)
            if st.button("🔄 " + _t("Udforsk alternativer", "Utforska alternativ", "Utforsk alternativer", "Etsi vaihtoehtoja", "Explore Alternatives"), use_container_width=True):
                if st.session_state.last_sector_prospects_list:
                    for p in st.session_state.last_sector_prospects_list:
                        st.session_state.prospect_ignore_list.append(p.get("symbol", "").upper())
                with st.spinner(_t("Søger efter alternative prospects...", "Söker efter alternativa prospekt...", "Søker efter alternative prospekter...", "Etsitään vaihtoehtoisia kohteita...", "Searching for alternatives...")):
                    prospect_json = generate_sector_prospects(GEMINI_API_KEY, selected_research_sector, st.session_state.user_name, st.session_state.prospect_ignore_list)
                    prospects_parsed = extract_json_array(prospect_json)
                    if prospects_parsed:
                        st.session_state.last_sector_prospects_list = prospects_parsed
                        st.session_state.last_research_sector_name = selected_research_sector
                        st.success(_t("Nye prospects klar!", "Nya prospekt redo!", "Nye prospekter klare!", "Uusia kohteita valmiina!", "New prospects ready!"))
                        time.sleep(1)
                        st.rerun()

        if st.button("🚀 " + _t("Engager rådgivere", "Engagera rådgivare", "Engasjer rådgivere", "Käynnistä neuvonantajat", "Engage Advisors"), use_container_width=True):
            st.session_state.prospect_ignore_list = []
            with st.spinner(_t("Rådgiverne analyserer sektoren og danner selskabs-prospekter...", "Rådgivarna analyserar sektorn och genererar prospekt...", "Rådgiverne analyserer sektoren og genererer prospekter...", "Neuvonantajat analysoivat sektoria ja luovat kohteita...", "The advisors are analyzing the sector and compiling investment cases...")):
                prospect_json = generate_sector_prospects(GEMINI_API_KEY, selected_research_sector, st.session_state.user_name, [])
                prospects_parsed = extract_json_array(prospect_json)
                if prospects_parsed:
                    st.session_state.last_sector_prospects_list = prospects_parsed
                    st.session_state.last_research_sector_name = selected_research_sector
                    st.rerun()

        # Visning af de fundne prospekter med interaktiv råds-debat karrusel (SWIPE/TABS PER CASE!)
        if st.session_state.last_sector_prospects_list:
            st.write(" ")
            st.subheader("📋 " + _t(f"Anbefalede prospects i {st.session_state.last_research_sector_name}", f"Rekommenderade innehav inom {st.session_state.last_research_sector_name}", f"Anbefalte prospects innen {st.session_state.last_research_sector_name}", f"Suositellut kohteet salkkuun {st.session_state.last_research_sector_name}", f"Recommended prospects in {st.session_state.last_research_sector_name}"))
            
            for idx, p in enumerate(st.session_state.last_sector_prospects_list):
                symbol = p.get("symbol", "Other").upper()
                name = p.get("name", "Other")
                thesis = p.get("thesis", "")
                opinions = p.get("opinions", {})
                
                with st.container(border=True):
                    col_info, col_btn = st.columns([3, 1])
                    with col_info:
                        # Tydelig visning af KID Handels-status
                        kid_badge = "🟢 " + _t("Handelsklar (Enkeltaktie - Intet KID påkrævet)", "Handelsklar (Enskild aktie)", "Handelsklar (Enkeltaksje)", "Kaupankäyntivalmis", "Tradeable (No KID needed)")
                        st.markdown(f"### {name} ({symbol})")
                        st.caption(kid_badge)
                        st.write(thesis)
                        
                        col_link1, col_link2, col_link3 = st.columns(3)
                        with col_link1:
                            st.markdown(f"[ Seeking Alpha]({p.get('sa_link')})")
                        with col_link2:
                            st.markdown(f"[ Yahoo Finance]({p.get('yf_link')})")
                        with col_link3:
                            st.markdown(f"[ Investor Relations]({p.get('ir_link')})")
                    
                    with col_btn:
                        st.write(" ")
                        st.write(" ")
                        if symbol in st.session_state.watchlist_list:
                            st.success(_t("Tilføjet! ", "Tillagd! ", "Lagt til! ", "Lisätty! ", "Added! "))
                        else:
                            if st.button(_t("➕ Watchlist", "➕ Watchlist", "➕ Watchlist", "➕ Tarkkailulista", "➕ Watchlist"), key=f"add_watchlist_{symbol}_{idx}", use_container_width=True):
                                st.session_state.watchlist_list.append(symbol)
                                st.success(_t(f"{symbol} tilføjet!", f"{symbol} tillagd!", f"{symbol} lagt til!", f"{symbol} lisätty!", f"{symbol} added!"))
                                time.sleep(1)
                                st.rerun()

                    # INTERAKTIV RÅDS-DEBAT SWIPE FOR DETTE ENKELTE PROSPEKT (SWOT MODEL!)
                    if opinions:
                        st.write(" ")
                        tab_names = [
                            "💪 " + _t("Styrker", "Styrkor", "Styrker", "Vahvuudet", "Strengths"), 
                            "⚠️ " + _t("Svagheder", "Svagheter", "Svakheter", "Heikkoudet", "Weaknesses"), 
                            "🚀 " + _t("Muligheder", "Möjligheter", "Muligheter", "Mahdollisuudet", "Opportunities"), 
                            "🛡️ " + _t("Trusler & Gæld", "Hot & Skulder", "Trusler & Gjeld", "Uhat & Velka", "Threats & Debt"), 
                            "🔵 " + _t("Saxo Forhold", "Saxo-villkor", "Saxo-forhold", "Saxo-ehdot", "Saxo Tradeability"), 
                            "👑 " + _t("Rådets Dom", "Rådets dom", "Rådets dom", "Neuvoston tuomio", "Strategic Verdict")
                        ]
                        prospect_tabs = st.tabs(tab_names)
                        with prospect_tabs[0]:
                            st.info(opinions.get("expansionist") or "No comment.")
                        with prospect_tabs[1]:
                            st.error(opinions.get("contrarian") or "No comment.")
                        with prospect_tabs[2]:
                            st.warning(opinions.get("outsider") or "No comment.")
                        with prospect_tabs[3]:
                            st.markdown(opinions.get("first_principles") or "No comment.")
                        with prospect_tabs[4]:
                            st.info(opinions.get("executor") or "No comment.")
                        with prospect_tabs[5]:
                            st.success(opinions.get("chairman") or "No comment.")

        # DYNAMISK, NYSKABENDE ETF PURIFIER (KORRIGERET TIL AT TAGE DYNAMISK TOP-N FRA BRUGEREN!)
        st.write("---")
        st.subheader("🗳️ " + _t("ETF Purifier", "ETF Purifier", "ETF Purifier", "ETF Purifier", "ETF Purifier"))
        st.write(_t(
            "Da du som privat investor ikke må købe brede amerikanske ETF'er på grund af KID-reglerne, kan du her indsætte en konventionel fond (f.eks. SPY eller QQQ). Vores AI-motor 'skræller' fonden, fjerner de ikke-kompatible selskaber, og præsenterer dig for de Shariah-godkendte enkeltaktier, du frit må handle enkeltvis på Saxo.",
            "Eftersom du som privat investerare inte får köpa breda amerikanska ETF:er på grund av KID-regler, kan du här klistra in en vanlig fond (t.ex. SPY eller QQQ). Vår AI 'skalar' fonden, rensar bort icke-kompatibla bolag och visar Shariah-godkända enskilda aktier som du fritt kan handla.",
            "Siden du som privat investor ikke kan kjøpe brede amerikanske ETF-er på grunn av KID-reglene, kan du her lime inn et vanlig fond (f.eks. SPY eller QQQ). Vår AI-motor 'skreller' fondet, fjerner de ikke-kompatible selskapene, og presenterer deg for de Shariah-godkjente enkeltaksjene du fritt kan handle.",
            "Koska et voi ostaa yhdysvaltalaisia ETF-rahastoja KID-sääntöjen vuoksi, voit syöttää tähän perinteisen rahaston (esim. SPY tai QQQ). Tekoälymme suodattaa rahaston, poistaa ei-yhteensopivat yhtiöt ja esittelee Shariah-hyväksytyt osakkeet, joita voit vapaasti ostaa Saxossa.",
            "Since you cannot buy broad US ETFs due to KID restrictions, enter a conventional ETF ticker (e.g., SPY or QQQ) below. Our AI 'peels' the ETF, purges non-compliant stocks, and presents Shariah-approved single stocks you can trade freely."
        ))
        
        col_peel_t, col_peel_n = st.columns([2, 1])
        with col_peel_t:
            peel_ticker = st.text_input(_t("Indtast konventionel ETF-ticker (f.eks. 'SPY', 'QQQ', 'MSCI'):", "Ange vanlig ETF-ticker:", "Skriv inn vanlig ETF-ticker:", "Syötä perinteinen ETF-ticker:", "Enter conventional ETF ticker (e.g., SPY, QQQ):"), placeholder="e.g., SPY")
        with col_peel_n:
            top_n = st.selectbox(
                _t("Antal selskaber at analysere:", "Antal bolag att analysera:", "Antall selskaper å analysere:", "Analysoitavien yritysten määrä:", "Top holdings to screen:"),
                [10, 20, 30, 40, 50],
                index=1
            )
        
        if st.button("🗳️ " + _t("Skræl ETF", "Skala ETF", "Skrell ETF", "Suodata ETF", "Peel ETF"), use_container_width=True):
            if peel_ticker:
                with st.spinner(_t("Scanner fondens største beholdninger og frasorterer uetisk gæld/forretning...", "Scannar fondens innehav...", "Scanner fondens posisjoner...", "Tarkistetaan rahaston omistuksia...", "Scanning holdings and purging non-compliant businesses...")):
                    # Sender det dynamiske valg (top_n) ind i purifier-motoren!
                    peeled_data = purify_conventional_etf(GEMINI_API_KEY, peel_ticker, top_n)
                    if peeled_data:
                        st.session_state.peeled_etf_list = peeled_data
                        st.session_state.last_peeled_etf_ticker = peel_ticker.upper()
                        st.rerun()

        if st.session_state.peeled_etf_list:
            st.write(" ")
            st.subheader("📋 " + _t(f"Resultater efter 'skrælning' af {st.session_state.last_peeled_etf_ticker}", f"Resultat efter skalning av {st.session_state.last_peeled_etf_ticker}", f"Resultater etter skrelling av {st.session_state.last_peeled_etf_ticker}", f"Hakutulokset suodatukselle {st.session_state.last_peeled_etf_ticker}", f"Purged & Screened results for {st.session_state.last_peeled_etf_ticker}"))
            
            for idx, item in enumerate(st.session_state.peeled_etf_list):
                ticker = item.get("symbol", "Other").upper()
                name = item.get("name", "Other")
                status = item.get("status", "Compliant")
                reason = item.get("reason", "")
                
                with st.container(border=True):
                    col_info, col_btn = st.columns([3, 1])
                    with col_info:
                        if status == "Compliant":
                            st.markdown(f"### 🟢 {name} ({ticker})")
                            st.caption("🟢 " + _t("Handelsklar (Enkeltaktie - Intet KID påkrævet)", "Handelsklar (Enskild aktie)", "Handelsklar (Enkeltaksje)", "Kaupankäyntivalmis", "Tradeable (No KID needed)"))
                        else:
                            st.markdown(f"### 🔴 ~~{name} ({ticker})~~")
                            st.caption("🔴 " + _t("Udelukket / Purged", "Utesluten / Purged", "Ekskludert / Purged", "Hylätty", "Purged"))
                        st.write(reason)
                    
                    with col_btn:
                        st.write(" ")
                        st.write(" ")
                        if status == "Compliant":
                            if ticker in st.session_state.watchlist_list:
                                st.success(_t("Tilføjet!", "Tillagd!", "Lagt til!", "Lisätty!", "Added!"))
                            else:
                                if st.button(_t("➕ Watchlist", "➕ Watchlist", "➕ Watchlist", "➕ Tarkkailulista", "➕ Watchlist"), key=f"add_peel_watchlist_{ticker}_{idx}", use_container_width=True):
                                    st.session_state.watchlist_list.append(ticker)
                                    st.success(_t(f"{ticker} tilføjet!", f"{ticker} tillagd!", f"{ticker} lagt til!", f"{ticker} lisätty!", f"{ticker} added!"))
                                    time.sleep(1)
                                    st.rerun()

        # Failsafe: Vis seneste ugentlige briefinger direkte på skærmen
        if st.session_state.generated_report:
            st.write("---")
            st.subheader("🗳️ " + _t("Dine seneste anbefalinger fra LLM Council", "Dina senaste rekommendationer från LLM Council", "Dine siste anbefalinger fra LLM Council", "Viimeisimmät LLM Council -suosituksesi", "Your Latest Recommendations"))
            
            if st.session_state.generated_audio_bytes:
                st.write("📢 " + _t("Lyt til din briefing her:", "Lyssna på din briefing här:", "Lytt til din briefing her:", "Kuuntele raporttisi tästä:", "Listen to your briefing here:"))
                st.audio(st.session_state.generated_audio_bytes, format="audio/mp3")
            
            with st.expander(_t("Læs hele rapporten på skærmen ▾", "Läs hela rapporten på skärmen ▾", "Les hele rapporten på skjermen ▾", "Lue koko raportti näytöllä ▾", "Read Full Report On-Screen ▾")):
                components.html(st.session_state.generated_report, height=500, scrolling=True)

    else:
        # NYSKREVET FORSIDE-TEKST FRA DIG (100% PRÆCIS INTEGRERET)
        render_html(_t("""
<div style="border: 1px solid #E2E8F0; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
    <h2 style="font-family: 'Georgia', serif; margin-top: 0; color: #0F172A !important;">🗳️ LLM Council</h2>
    <h3 style="font-family: 'Georgia', serif; margin-top: 0; color: #C5A880 !important;">🛡️ Den nemme vej til en global og Shariah-kompatibel portefølje</h3>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        Det burde være nemt at opbygge en rolig, passiv og etisk formue. Men som muslimsk investor i Norden rammer man hurtigt en mur.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        På grund af komplekse EU-regler er mange store islamiske fonde spærret på de nordiske handelsplatforme. Det efterlader dig i en umulig blindgyde: Enten skal du gå på kompromis med dine værdier, lade helt være med at investere, eller bruge uoverskueligt meget tid på selv at lege aktieanalytiker og scanne regnskaber.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        <strong>LLM Council er sat i verden for at åbne markedet op igen.</strong>
    </p>
    <p style="margin-bottom: 0; font-size: 15px; line-height: 1.6;">
        Vi finkæmmer det globale marked for dig – fra brede islamiske indeksfonde (som Invesco Dow Jones) og spændende regions-ETF'er (som f.eks. Saudi-Arabien) til stærke enkeltaktier. Vores værktøj sorterer automatisk de låste roadblocks fra, så du kun præsenteres for Shariah-godkendte investeringer, du rent faktisk kan købe direkte fra din foretrukne nordiske platform.
    </p>
</div>
""", """
<div style="border: 1px solid #E2E8F0; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
    <h2 style="font-family: 'Georgia', serif; margin-top: 0; color: #0F172A !important;">🗳️ LLM Council</h2>
    <h3 style="font-family: 'Georgia', serif; margin-top: 0; color: #C5A880 !important;">🛡️ Den enkla vägen till en global och Shariah-kompatibel portfölj</h3>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        Det borde vara enkelt att bygga en lugn, passiv och etisk förmögenhet. Men som muslimsk investerare i Norden slår man snabbt i en vägg.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        På grund av komplexa EU-regler är många stora islamiska fonder spärrade på de nordiska handelsplattformarna. Det lämnar dig i en omöjlig återvändsgränd: Antingen måste du kompromissa med dina värderingar, helt avstå från att investera, eller lägga orimligt mycket tid på att själv leka aktieanalytiker och scanna årsredovisningar.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        <strong>LLM Council skapades för att öppna marknaden igen.</strong>
    </p>
    <p style="margin-bottom: 0; font-size: 15px; line-height: 1.6;">
        Vi finkämme den globala marknaden åt dig – från breda islamiska indexfonde (som Invesco Dow Jones) och spännande region-ETF:er (som f.ex. Saudiarabien) till starka enskilda aktier. Vårt verktyg sorterar automatiskt bort de låsta hindren, så att du bara presenteras för Shariah-godkända investeringar som du faktiskt kan köpa direkt från din favoritplattform i Norden.
    </p>
</div>
""", """
<div style="border: 1px solid #E2E8F0; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
    <h2 style="font-family: 'Georgia', serif; margin-top: 0; color: #0F172A !important;">🗳️ LLM Council</h2>
    <h3 style="font-family: 'Georgia', serif; margin-top: 0; color: #C5A880 !important;">🛡️ Den nemme veien til en global og Shariah-kompatibel portefølje</h3>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        Det burde være enkelt å bygge en rolig, passiv og etisk formue. Men som muslimsk investor i Norden rammer man raskt en mur.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        På grunn av komplekse EU-regler er mange store islamiske fond sperret på de nordiske handelsplattformene. Det etterlater deg i en umulig blindvei: Enten må du inngå kompromiss med dine verdier, la helt være å investere, eller bruke uoverskuelig mye tid på selv å leke aksjeanalytiker og scanne regnskaper.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        <strong>LLM Council er satt i verden for å åpne markedet opp igen.</strong>
    </p>
    <p style="margin-bottom: 0; font-size: 15px; line-height: 1.6;">
        We scan the global market for you – from broad Islamic index funds (like Invesco Dow Jones) and exciting region-ETFs (such as Saudi Arabia) to strong single stocks. Our tool automatically filters out the locked roadblocks, so you are only presented with Shariah-approved investments you can actually buy directly from your favorite Nordic platform.
    </p>
</div>
""", """
<div style="border: 1px solid #E2E8F0; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
    <h2 style="font-family: 'Georgia', serif; margin-top: 0; color: #0F172A !important;">🗳️ LLM Council</h2>
    <h3 style="font-family: 'Georgia', serif; margin-top: 0; color: #C5A880 !important;">🛡️ Helppo tie globaaliin ja Shariah-yhteensopivaan salkkuun</h3>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        Rauhallisen, passiivisen ja eettisen varallisuuden kerryttämisen pitäisi olla helppoa. Mutta muslimisijoittajana Pohjoismaissa törmäät nopeasti seinään.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        Monimutkaisten EU-sääntöjen vuoksi monet suuret islamilaiset rahastot on estetty pohjoismaisissa kaupankäyntialustoissa. Tämä jättää sinut mahdottomaan umpikujaan: sinun on joko tingittävä arvoistasi, jätettävä sijoittamatta kokonaan tai käytettävä kohtuuttomasti aikaa osakeanalyytikon leikkimiseen ja tilinpäätösten suodattamiseen.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        <strong>LLM Council on perustettu avaamaan markkinat sinulle uudelleen.</strong>
    </p>
    <p style="margin-bottom: 0; font-size: 15px; line-height: 1.6;">
        Kartoitamme globaalit markkinat puolestasi – laajoista islamilaisista indeksirahastoista (kuten Invesco Dow Jones) ja mielenkiintoisista alueellisista ETF-rahastoista (kuten Saudi-Arabia) vahvoihin yksittäisiin osakkeisiin. Työkalumme suodattaa lukitut esteet automaattisesti pois, jotta sinulle esitellään vain Shariah-hyväksyttyjä sijoituksia, joita voit aidosti ostaa suoraan pohjoismaisesta alustastasi.
    </p>
</div>
""", """
<div style="border: 1px solid #E2E8F0; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
    <h2 style="font-family: 'Georgia', serif; margin-top: 0; color: #0F172A !important;">🗳️ LLM Council</h2>
    <h3 style="font-family: 'Georgia', serif; margin-top: 0; color: #C5A880 !important;">🛡️ The Easy Way to a Global, Shariah-Compliant Portfolio</h3>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        Building a calm, passive, and ethical wealth journey should be easy. Yet, as a Muslim investor in the Nordics, you quickly run into a wall.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        Due to complex EU regulations, many large Islamic funds are blocked on Nordic brokerage platforms. This leaves you in an impossible dead-end: you must either compromise on your values, abstain from investing altogether, or spend endless hours playing stock analyst and auditing financial statements.
    </p>
    <p style="margin-bottom: 12px; font-size: 15px; line-height: 1.6;">
        <strong>LLM Council was created to reopen the market for you.</strong>
    </p>
    <p style="margin-bottom: 0; font-size: 15px; line-height: 1.6;">
        We scan the global market for you – from broad Islamic index funds (like Invesco Dow Jones) and exciting regional ETFs (such as Saudi Arabia) to strong individual stocks. Our tool automatically filters out the locked roadblocks, so you are only presented with Shariah-approved investments you can actually buy directly from your favorite Nordic platform.
    </p>
</div>
"""))
        
        # Stor velkomst-knap
        start_btn_text = _t("Opsæt dit porteføljestyringsværktøj ➔", "Konfigurera ditt portföljverktyg ➔", "Opprett ditt porteføljestyringsverktøy ➔", "Määritä salkunhallintatyökalusi ➔", "Setup your portfolio management tool ➔")
        if st.button(start_btn_text, use_container_width=True):
            st.session_state.step = 2
            st.rerun()
            
        st.write("---")
        
        # LOGIND FOR EKSISTERENDE BRUGERE PÅ VELKOMSTSIDEN
        st.subheader("🔑 " + _t("Eksisterende bruger? Log ind her", "Befintlig användare? Logga in här", "Eksisterende bruker? Logg inn her", "Nykyinen käyttäjä? Kirjaudu sisään tästä", "Existing user? Log in here"))
        login_email = st.text_input("E-mail", placeholder="navn@gmail.com", key="home_login_email")
        login_password = st.text_input(_t("Adgangskode", "Lösenord", "Adgangskode", "Salasana", "Password"), type="password", key="home_login_password")
        
        if st.button(_t("Log ind ➔", "Logga in ➔", "Logg inn ➔", "Kirjaudu sisään ➔", "Log In ➔"), use_container_width=True, key="home_login_btn"):
            if login_email and "@" in login_email and login_password:
                with st.spinner(_t("Forbinder til profil...", "Ansluter...", "Forbinder...", "Yhdistetään...", "Connecting...")):
                    response = requests.get(f"{DATABASE_URL}?email={login_email}&password={login_password}", timeout=10)
                    if response.status_code == 200:
                        res_data = response.json()
                        if res_data.get("status") == "success":
                            db_profile = res_data
                            st.session_state.investor_holdings = db_profile.get("holdings", [])
                            st.session_state.targets = db_profile.get("targets", {"Aktier": 25.0, "Sukuk": 25.0, "Råvarer": 25.0, "Kontanter/Private": 25.0})
                            st.session_state.slider_stocks = int(st.session_state.targets.get("Aktier", 25.0))
                            st.session_state.slider_sukuk = int(st.session_state.targets.get("Sukuk", 25.0))
                            st.session_state.slider_commodities = int(st.session_state.targets.get("Råvarer", 25.0))
                            st.session_state.slider_cash = int(st.session_state.targets.get("Kontanter/Private", 25.0))
                            st.session_state.horizon = db_profile.get("horizon", "7-15 years")
                            st.session_state.user_name = db_profile.get("name", "Investor")
                            st.session_state.frequency = db_profile.get("frequency", "Weekly")
                            st.session_state.user_email = login_email
                            st.session_state.is_logged_in = True
                            st.success(_t("Log ind fuldført!", "Inloggad!", "Logg inn fullført!", "Kirjautuminen onnistui!", "Log in successful!"))
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error(_t("Forkert login eller adgangskode.", "Felaktig e-post eller lösenord.", "Feil login eller adgangskode.", "Väärä sähköpostiosoite tai salasana.", "Incorrect login or password."))


# --- TRIN 2: LOGIN & PROFILOPRETTELSE ---
elif st.session_state.step == 2:
    st.subheader(_t("Opret din profil eller log ind", "Skapa din profil eller logga in", "Opprett din profil eller logg inn", "Luo profiili tai kirjaudu sisään", "Create your profile or sign in"))
    st.write(_t("Dine oplysninger gemmes sikkert, så din personlige portefølje og dine ugentlige briefinger synkroniseres automatisk.", "Dina uppgifter sparas säkert så att din personliga portfölj synkroniseras automatiskt.", "Dine opplysninger gemmes sikkert slik at din personlige portefølje synkroniseres automatisk.", "Tietosi tallennetaan turvallisesti salkun synkronointia varten.", "Your details are stored securely so your personal portfolio syncs automatically."))
    
    col_l1, col_l2 = st.columns(2)
    with col_l1:
        login_email = st.text_input("E-mail", placeholder="navn@gmail.com", value=st.session_state.user_email, key="step2_login_email")
    with col_l2:
        login_password = st.text_input(_t("Adgangskode", "Lösenord", "Adgangskode", "Salasana", "Password"), type="password", key="step2_login_password")

    is_new_user = False
    db_profile = None

    if st.button(_t("Log ind ➔", "Logga in ➔", "Logg inn ➔", "Kirjaudu sisään ➔", "Log In ➔"), use_container_width=True, key="step2_login_btn"):
        if login_email and "@" in login_email and login_password:
            with st.spinner(_t("Forbinder til profil...", "Ansluter...", "Forbinder...", "Yhdistetään...", "Connecting...")):
                response = requests.get(f"{DATABASE_URL}?email={login_email}&password={login_password}", timeout=10)
                if response.status_code == 200:
                    res_data = response.json()
                    if res_data.get("status") == "success":
                        db_profile = res_data
                        st.session_state.investor_holdings = db_profile.get("holdings", [])
                        st.session_state.targets = db_profile.get("targets", {"Aktier": 25.0, "Sukuk": 25.0, "Råvarer": 25.0, "Kontanter/Private": 25.0})
                        st.session_state.slider_stocks = int(st.session_state.targets.get("Aktier", 25.0))
                        st.session_state.slider_sukuk = int(st.session_state.targets.get("Sukuk", 25.0))
                        st.session_state.slider_commodities = int(st.session_state.targets.get("Råvarer", 25.0))
                        st.session_state.slider_cash = int(st.session_state.targets.get("Kontanter/Private", 25.0))
                        st.session_state.horizon = db_profile.get("horizon", "7-15 years")
                        st.session_state.user_name = db_profile.get("name", "Investor")
                        st.session_state.frequency = db_profile.get("frequency", "Weekly")
                        st.session_state.user_email = login_email
                        st.session_state.is_logged_in = True
                        st.success(_t("Velkommen tilbage!", "Välkommen tillbaka!", "Velkommen tilbake!", "Tervetuloa takaisin!", "Welcome back!"))
                        time.sleep(1)
                        st.rerun()
                    elif res_data.get("status") == "incorrect_password":
                        st.error(_t("Forkert adgangskode.", "Felaktigt lösenord.", "Feil adgangskode.", "Väärä salasana.", "Incorrect password."))
                    elif res_data.get("status") == "not_found":
                        is_new_user = True

    if is_new_user:
        st.info(_t("E-mailen blev ikke fundet. Opret en profil herunder:", "E-postadressen hittades inte. Skapa en profil nedan:", "E-mailen ble ikke funnet. Opprett en profil nedenfor:", "Sähköpostiosoitetta ei löytynyt. Luo profiili alta:", "Email not found. Create a profile below:"))
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            confirm_password = st.text_input(_t("Bekræft adgangskode", "Bekräfta lösenord", "Bekreft adgangskode", "Vahvista salasana", "Confirm password"), type="password")
        with col_s2:
            signup_name = st.text_input(_t("Dit fulde navn", "Ditt fullständiga namn", "Ditt fulle navn", "Koko nimesi", "Your full name"), value="Investor")
        
        if st.button(_t("📝 Registrer profil", "📝 Skapa profil", "📝 Registrer profil", "📝 Rekisteröidy", "📝 Register profile"), use_container_width=True):
            if login_password != confirm_password:
                st.error(_t("Adgangskoderne matcher ikke!", "Lösenorden matchar inte!", "Adgangskodene matcher ikke!", "Salasanat eivät täsmää!", "Passwords do not match!"))
            elif not signup_name:
                st.error(_t("Udfyld venligst dit navn.", "Vänligen fyll i ditt namn.", "Vennligst oppgi navnet ditt.", "Anna nimesi.", "Please enter your name."))
            else:
                status = save_user_portfolio_to_db(
                    email=login_email,
                    password=login_password,
                    holdings=st.session_state.investor_holdings,
                    targets=st.session_state.targets,
                    horizon=st.session_state.horizon,
                    name=signup_name,
                    frequency=st.session_state.frequency
                )
                if status == "success":
                    st.session_state.user_name = signup_name
                    st.session_state.user_email = login_email
                    st.session_state.is_logged_in = True
                    st.success(_t("Profilen blev oprettet!", "Profilen skapad!", "Profilen ble opprettet!", "Profiili luotu!", "Profile created!"))
                    st.session_state.step = 3
                    st.rerun()

    # Navigationsknapper
    st.write(" ")
    col_prev, col_next = st.columns(2)
    with col_prev:
        if st.button("⬅ " + _t("Tilbage", "Tillbaka", "Tilbake", "Takaisin", "Back"), use_container_width=True, key="step2_prev_btn_normal"):
            st.session_state.step = 1
            st.rerun()
    with col_next:
        if st.session_state.is_logged_in:
            if st.button(_t("Næste trin", "Nästa steg", "Neste trinn", "Seuraava vaihe", "Next step") + " ➔", use_container_width=True):
                st.session_state.step = 3
                st.rerun()


# --- TRIN 3: INVESTERINGSPROFIL & ALLOKERING (INTELIGENTE SLIDERE UDEN OVER-SHOOT) ---
elif st.session_state.step == 3:
    st.subheader(_t("Definer din investeringsprofil", "Definiera din investeringsprofil", "Definer din investeringsprofil", "Määritä sijoitusprofiilisi", "Define your investment profile"))
    
    col_n1, col_n2 = st.columns(2)
    with col_n1:
        st.session_state.user_name = st.text_input(_t("Dit navn i rapporten:", "Ditt namn i rapporten:", "Ditt namn i rapporten:", "Nimesi raportissa:", "Your name in the report:"), value=st.session_state.user_name)
    with col_n2:
        st.session_state.user_email = st.text_input(_t("E-mailadresse til briefinger:", "E-postadress för briefinger:", "E-postadresse til briefinger:", "Sähköpostiosoite briefingeille:", "Email address for briefings:"), value=st.session_state.user_email)

    col_s1, col_s2 = st.columns(2)
    with col_s1:
        horizon_options = ["1-3 years", "3-7 years", "7-15 years", "15+ years"]
        horizon_index = horizon_options.index(st.session_state.horizon) if st.session_state.horizon in horizon_options else 2
        st.session_state.horizon = st.selectbox(_t("Investeringshorisont:", "Placeringshorisont:", "Investeringshorisont:", "Sijoitusaika:", "Investment horizon:"), horizon_options, index=horizon_index)
    with col_s2:
        freq_options = ["Daily", "Weekly", "Bi-weekly", "Monthly"]
        freq_index = freq_options.index(st.session_state.frequency) if st.session_state.frequency in freq_options else 1
        st.session_state.frequency = st.selectbox(_t("Hvor ofte ønsker du briefing?", "Hur ofta vill du ha briefing?", "Hvor ofte ønsker du briefing?", "Kuinka usein haluat raportin?", "How often do you want briefings?"), freq_options, index=freq_index)

    # REBALANCERINGS BUDGET INDTASTNING TIL PORTEFØLJEPLEJE
    st.write("---")
    st.subheader("💰 " + _t("Dit løbende rebalancerings-budget", "Månatlig sparbudget", "Ditt løpende rebalanceringsbudsjett", "Kuukausisäästöbudjetti", "Your Planned Investment Budget"))
    st.session_state.monthly_deposit = st.number_input(
        _t("Planlagt månedlig opsparing (DKK):", "Månatligt sparande (DKK):", "Planlagt månedlig sparing (DKK):", "Suunniteltu kuukausisäästö (DKK):", "Planned monthly savings (DKK):"),
        min_value=0,
        value=int(st.session_state.monthly_deposit)
    )

    st.write("---")
    st.subheader(_t("Angiv din ønskede mål-allokering", "Ange din önskade måstallokering", "Angi din ønskede målallokering", "Aseta tavoitesalkkusi hajautus", "Specify your target asset allocation"))
    st.write(_t("Træk i sliderne nedenfor. Din samlede vægtning låses automatisk, så den ALDRIG kan skyde over 100% samlet.", "Dra i reglagen nedan. Din totala allokering låses automatiskt so att den ALDRIG kan överstiga 100% totalt.", "Dra i sliderne nedenfor. Din samlede vekting låses automatisk, slik at den ALDRIG kan overskride 100% samlet.", "Säädä liukusäätimiä alta. Tavoiteosuus lukitaan automaattisesti siten, että kokonaishajautus ei voi ylittää 100 %.", "Adjust the sliders below. Your total allocation is automatically capped, so it can NEVER exceed 100% in total."))
    
    # Loft-beregninger (zero-sum budget)
    st.session_state.slider_stocks = min(st.session_state.slider_stocks, 100)
    max_sukuk = 100 - st.session_state.slider_stocks
    st.session_state.slider_sukuk = min(st.session_state.slider_sukuk, max_sukuk)
    max_commodities = 100 - st.session_state.slider_stocks - st.session_state.slider_sukuk
    st.session_state.slider_commodities = min(st.session_state.slider_commodities, max_commodities)
    max_cash = 100 - st.session_state.slider_stocks - st.session_state.slider_sukuk - st.session_state.slider_commodities
    st.session_state.slider_cash = min(st.session_state.slider_cash, max_cash)

    # Slidere med dynamisk max-værdi
    target_stocks = st.slider(_t("Equities (Aktier) %", "Equities (Aktier) %", "Equities (Aksjer) %", "Osakkeet %", "Equities %"), 0, 100, key="slider_stocks")
    target_sukuk = st.slider(_t("Sukuk %", "Sukuk %", "Sukuk %", "Sukuk %", "Sukuk %"), 0, max_sukuk, key="slider_sukuk")
    target_commodities = st.slider(_t("Commodities (Råvarer) %", "Commodities (Råvarer) %", "Commodities (Råvarer) %", "Raaka-aineet %", "Commodities %"), 0, max_commodities, key="slider_commodities")
    target_cash = st.slider(_t("Cash/Private %", "Cash/Private %", "Cash/Private %", "Käteinen/Yksityinen %", "Cash/Private %"), 0, max_cash, key="slider_cash")

    total_target = target_stocks + target_sukuk + target_commodities + target_cash
    
    if total_target != 100:
        difference = 100 - total_target
        st.warning(_t(
            f"⚠️ Allokeringen skal give 100% tilsammen. Nuværende sum: {total_target}%. Du mangler at fordele {difference}%.",
            f"⚠️ Allokeringen måste bli totalt 100%. Nuvarande summa: {total_target}%. Du behöver fördela {difference}%.",
            f"⚠️ Allokeringen må gi 100% til sammen. Nåværende sum: {total_target}%. Du mangler å fordele {difference}%.",
            f"⚠️ Hajautuksen on oltava tasan 100 %. Nykyinen summa: {total_target} %. Sinun on jaettava vielä {difference} %.",
            f"⚠️ Allocation must equal 100% in total. Current sum: {total_target}%. You need to allocate {difference}%."
        ))
    else:
        st.success(_t("✅ Allokeringen er præcis 100%! Du kan nu fortsætte.", "✅ Allokeringen är exakt 100%! Du kan nu gå vidare.", "✅ Allocation is exactly 100%! Du kan nu gå videre.", "✅ Hajautus on tasan 100 %! Voit jatkaa eteenpäin.", "✅ Allocation is exactly 100%! You can now proceed."))
        st.session_state.targets = {
            "Aktier": float(target_stocks),
            "Sukuk": float(target_sukuk),
            "Råvarer": float(target_commodities),
            "Kontanter/Private": float(target_cash)
        }

    # Navigationsknapper
    st.write(" ")
    col_prev, col_next = st.columns(2)
    with col_prev:
        if st.button("⬅ " + _t("Tilbage", "Tillbaka", "Tilbake", "Takaisin", "Back"), use_container_width=True, key="step3_prev_btn"):
            st.session_state.step = 2
            st.rerun()
    with col_next:
        st.button(
            _t("Næste trin", "Nästa steg", "Neste trinn", "Seuraava vaihe", "Next step") + " ➔", 
            use_container_width=True, 
            disabled=(total_target != 100),
            key="next_to_step4"
        )
        if st.session_state.get("next_to_step4"):
            st.session_state.step = 4
            st.rerun()


# --- TRIN 4: PORTEFØLJEOPBYGNING (NYSKABENDE ETF-KLONING INDBYGGET) ---
elif st.session_state.step == 4:
    st.subheader(_t("Indtast dine nuværende aktiver", "Fyll i dina nuvarande tillgångar", "Oppgi dine nåværende aktiver", "Syötä nykyiset sijoituksesi", "Input your current holdings"))
    
    # Gemmes nu stabilt i session-state så det huskes mellem trin
    st.session_state.is_new_investor = st.checkbox(
        _t("Jeg er helt ny investor (starter fra bunden med tom portefølje)", "Jag är en helt ny investerare (börjar från början med tom portfölj)", "Jeg er helt ny investor (starter fra bunnen med tom portefølje)", "Olen täysin uusi sijoittaja (aloitan tyhjästä tyhjällä salkulla)", "I am a completely new investor (starting from scratch with an empty portfolio)"),
        value=st.session_state.is_new_investor
    )

    selected_new_sectors = []
    if st.session_state.is_new_investor:
        
        # 1. OPTION TIL DIREKTE KLONING AF LÅSTE AMERIKANSKE FONDER (NYSKABENDE!)
        st.write("---")
        st.write("🛡️ **" + _t("Genvej: Klon en låst Shariah-ETF med enkeltaktier", "Genväg: Klona en låst Shariah-ETF med enskilda aktier", "Snarvei: Klon en låst Shariah-ETF med enkeltaksjer", "Pika-asetus: Kloonaa lukittu Shariah-ETF yksittäisillä osakkeilla", "Shortcut: Clone a locked Shariah-ETF with individual stocks") + "**")
        st.write(_t(
            "Da amerikanske fonde (som HLAL) mangler KID og er låst på Saxo Bank, kan du her med ét klik klone fondens 5 største selskaber direkte ind i din portefølje:",
            "Eftersom amerikanska fonder (som HLAL) saknar KID och är låsta på Saxo Bank, kan du här med ett klick klona fondens 5 största innehav direkt i din portfölj:",
            "Siden amerikanske fond (som HLAL) mangler KID og er låst på Saxo Bank, kan du her med ett klikk klone fondets 5 største posisjoner direkte inn i din portefølje:",
            "Koska yhdysvaltalaisilta rahastoilta (kuten HLAL) puuttuu KID-asiakirja ja ne on lukittu Saxo Bankissa, voit tästä kloonata rahaston 5 suurinta omistusta yhdellä klikkauksella suoraan salkkuusi:",
            "Since US-domiciled ETFs (like HLAL) lack KIDs and are locked on Saxo, you can clone the ETF's top 5 holdings directly into your portfolio with a single click:"
        ))
        if st.button("🗳️ " + _t("Klon den globale Shariah ETF (HLAL)", "Klona Shariah-ETF:en (HLAL)", "Klon den globale Shariah ETF (HLAL)", "Kloonaa Shariah ETF (HLAL)", "Clone global Shariah ETF (HLAL)"), use_container_width=True):
            st.session_state.investor_holdings = STATIC_ETF_CLONES["HLAL"].copy()
            st.session_state.is_new_investor = False
            st.success(_t("Boom! Fondens selskaber (MSFT, AAPL, NVDA, GOOGL, CRM) er nu tilføjet. Du er ikke længere tom!", "Boom! Fondens bolag har lagts till. Du är inte längre tom!", "Boom! Fondets selskaper er nå lagt til. Du er ikke lenger tom!", "Hienoa! Rahaston osakkeet on nyt lisätty salkkuusi!", "Success! The ETF's holdings have been added. Your portfolio is ready!"))
            time.sleep(1.5)
            st.rerun()

        st.write("---")
        with st.expander(_t("Eller vælg de sektorer du vil opbygge eksponering mod ▾", "Eller välj de sektorer du vill bygga exponering mot ▾", "Eller velg sektorene du vil bygge eksponering mot ▾", "Tai valitse salkkusi kohdesektorit ▾", "Or select the sectors you want to build exposure to ▾")):
            for sector in TARGET_SUBSECTORS:
                if st.checkbox(sector, value=False):
                    selected_new_sectors.append(sector)
    else:
        is_manual = st.checkbox(_t("Er dette et manuelt aktiv? (F.eks. kontantbeholdning, unoterede selskaber)", "Är detta en manuell tillgång? (T.ex. kontanter, onoterade bolag)", "Er dette et manuelt aktiv? (F.eks. kontantbeholdning, unoterte selskaper)", "Onko tämä manuaalinen omaisuuserä? (Esim. käteinen, listaamattomat yhtiöt)", "Is this a manual asset? (e.g., cash, private equity)"))

        if is_manual:
            col_m1, col_m2 = st.columns(2)
            with col_m1:
                manual_name = st.text_input(_t("Navn på aktiv:", "Namn på tillgång:", "Navn på aktiv:", "Omaisuuserän nimi:", "Asset name:"), placeholder="F.eks. Saxo Kontant DKK")
            with col_m2:
                manual_value = st.number_input(_t("Samlet værdi i DKK:", "Total värde i DKK:", "Samlet verdi i DKK:", "Kokonaisarvo DKK:", "Total value in DKK:"), min_value=1, value=1000)
                
            col_m3, col_m4 = st.columns(2)
            with col_m3:
                manual_category = st.selectbox(_t("Aktivklasse:", "Tillgångsslag:", "Aktivklasse:", "Omaisuusluokka:", "Asset class:"), ["Cash/Private", "Sukuk", "Commodities"])
            with col_m4 = st.columns(1):
                manual_sector = st.selectbox(_t("Delsektor:", "Delsektor:", "Delsektor:", "Sektori:", "Sub-sector:"), TARGET_SUBSECTORS + ["Kontanter", "Private investeringer"])
                
            if st.button(_t("➕ Tilføj manuelt aktiv", "➕ Lägg till manuell tillgång", "➕ Legg til manuelt aktiv", "➕ Lisää manuaalinen omaisuuserä", "➕ Add manual asset"), use_container_width=True):
                if manual_name:
                    virtual_ticker = f"PVT_{manual_name.upper().replace(' ', '_')}"
                    st.session_state.investor_holdings.append({
                        "Company Name": manual_name,
                        "Ticker": virtual_ticker,
                        "Shares": 1,
                        "Category": UI_TO_DB_MAP.get(manual_category, "Kontanter/Private"),
                        "Sector": manual_sector,
                        "manual_value": manual_value,
                        "Kurs": manual_value
                    })
                    st.success(_t(f"Tilføjede {manual_name}.", f"La till {manual_name}.", f"La til {manual_name}.", f"Lisätty {manual_name}.", f"Added {manual_name}."))
                    time.sleep(1)
                    st.rerun()
        else:
            search_query = st.text_input(_t("🔍 Søg efter selskab eller ticker:", "🔍 Sök efter bolag eller ticker:", "🔍 Søk etter selskap eller ticker:", "🔍 Etsi yritystä tai tickeriä:", "🔍 Search by company name or ticker:"))

            if search_query:
                search_results = search_tickers_by_name_multi(search_query)
                if search_results:
                    options_format = [f"{r['name']} ({r['symbol']})" for r in search_results]
                    selected_option_str = st.selectbox(_t("Vælg det rigtige aktiv fra listen:", "Välj rätt tillgång från listan:", "Velg riktig aktiv fra listen:", "Valitse oikea sijoituskohde:", "Select the correct asset:"), options_format)
                    
                    selected_idx = options_format.index(selected_option_str)
                    target_asset = search_results[selected_idx]
                    resolved_ticker = target_asset["symbol"]
                    comp_name = target_asset["name"]
                    
                    try:
                        cat, sub_sec = get_category_and_sector_failsafe(resolved_ticker, target_category=st.session_state.targets)
                        display_cat = DB_TO_UI_MAP.get(cat, cat)
                        
                        render_html(f"""
<div style="background-color: #F8FAFC; border: 1px solid #C5A880; padding: 15px; border-radius: 6px; margin-top: 15px; margin-bottom: 15px;">
    <strong>🔍 Match:</strong> {comp_name} ({resolved_ticker})<br>
    <strong>Kategori:</strong> {display_cat} | <strong>Sektor:</strong> {sub_sec}
</div>
""")
                        
                        col_shares, col_add = st.columns([1, 1])
                        with col_shares:
                            shares_to_add = st.number_input(_t("Antal aktier ejet:", "Antal aktier ägda:", "Antall aksjer ejet:", "Omistettujen osakkeiden määrä:", "Shares owned:"), min_value=1, value=10)
                        with col_add:
                            st.write(" ")
                            st.write(" ")
                            if st.button(_t("➕ Tilføj til min portefølje", "➕ Lägg till i portfölj", "➕ Legg til i min portefølje", "➕ Lisää salkkuun", "➕ Add to portfolio"), use_container_width=True):
                                exists = False
                                for h in st.session_state.investor_holdings:
                                    if h["Ticker"] == resolved_ticker:
                                        h["Shares"] += shares_to_add
                                        exists = True
                                        break
                                if not exists:
                                    st.session_state.investor_holdings.append({
                                        "Company Name": comp_name,
                                        "Ticker": resolved_ticker,
                                        "Shares": shares_to_add,
                                        "Category": cat,
                                        "Sector": sub_sec,
                                        "Kurs": 0.0
                                    })
                                st.success(_t(f"Tilføjede {shares_to_add} stk. {comp_name}.", f"La till {shares_to_add} st. {comp_name}.", f"La till {shares_to_add} stk. {comp_name}.", f"Lisätty {shares_to_add} kpl {comp_name}.", f"Added {shares_to_add} shares of {comp_name}."))
                                time.sleep(1)
                                st.rerun()
                    except Exception as e:
                        st.error(f"Kunne ikke hente data for {resolved_ticker}: {str(e)}")

    st.write("---")
    st.write(_t("### Dine aktive positioner:", "### Dina aktiva positioner:", "### Dine aktive posisjoner:", "### Aktiiviset sijoituksesi:", "### Your active holdings:"))
    if st.session_state.investor_holdings and not st.session_state.is_new_investor:
        holdings_df = pd.DataFrame(st.session_state.investor_holdings)
        holdings_df['Category_Display'] = holdings_df['Category'].apply(lambda x: DB_TO_UI_MAP.get(x, x))
        
        edited_holdings = st.data_editor(
            holdings_df,
            num_rows="dynamic",
            column_config={
                "Company Name": st.column_config.TextColumn(_t("Navn", "Namn", "Navn", "Nimi", "Name"), disabled=True),
                "Ticker": st.column_config.TextColumn("Ticker", disabled=True),
                "Shares": st.column_config.NumberColumn(_t("Antal", "Antal", "Antall", "Määrä", "Shares"), min_value=1),
                "Category_Display": st.column_config.TextColumn(_t("Aktivklasse", "Tillgångsslag", "Aktivklasse", "Omaisuusluokka", "Asset class"), disabled=True),
                "Sector": st.column_config.TextColumn(_t("Delsektor", "Delsektor", "Delsektor", "Sektori", "Sub-sector"), disabled=True),
                "manual_value": st.column_config.NumberColumn(_t("Manuel værdi (DKK)", "Manuellt värde (DKK)", "Manuell verdi (DKK)", "Manuaalinen arvo (DKK)", "Manual value (DKK)"), min_value=0)
            },
            use_container_width=True,
            key="holdings_editor"
        )
        
        if not edited_holdings.equals(holdings_df):
            st.session_state.investor_holdings = edited_holdings.to_dict(orient="records")
            st.rerun()
            
        if st.button(_t("💾 Gem ændringer i min profil", "💾 Spara ändringar i profil", "💾 Lagre endringer i profil", "💾 Tallenna profiiliin", "💾 Save changes to my profile"), use_container_width=True):
            status = save_user_portfolio_to_db(
                email=st.session_state.user_email,
                password=login_password,
                holdings=st.session_state.investor_holdings,
                targets=st.session_state.targets,
                horizon=st.session_state.horizon,
                name=st.session_state.user_name,
                frequency=st.session_state.frequency
            )
            if status == "success":
                st.success(_t("Ændringerne blev gemt på din profil!", "Ändringarna har sparats!", "Endringene ble lagret!", "Tallennettu onnistuneesti!", "Changes saved successfully!"))
    elif st.session_state.is_new_investor:
        st.info(_t("Nybegynder-status aktiveret. Du behøver ikke indtaste beholdninger.", "Nybörjarstatus aktiverad. Inga innehav krävs.", "Nybegynner-status aktivert. Du trenger ikke oppgi beholdninger.", "Aloittelija-tila aktivoitu. Salkun osia ei tarvitse syöttää.", "New investor status activated. No holdings required."))
    else:
        st.info(_t("Porteføljen er tom lige nu. Tilføj aktiver herover for at komme videre.", "Portföljen är tom. Lägg till tillgångar ovan för att gå vidare.", "Porteføljen er tom. Legg til aktiver ovenfor for å fortsette.", "Salkku on tyhjä. Lisää sijoituksia ylhäältä jatkaaksesi.", "Your portfolio is empty. Add assets above to proceed."))

    # Navigationsknapper
    st.write(" ")
    col_prev, col_next = st.columns(2)
    with col_prev:
        if st.button("⬅ " + _t("Tilbage", "Tillbaka", "Tilbake", "Takaisin", "Back"), use_container_width=True, key="step4_prev_btn"):
            st.session_state.step = 3
            st.rerun()
    with col_next:
        if st.button(_t("Næste trin", "Nästa steg", "Neste trinn", "Seuraava vaihe", "Next step") + " ➔", use_container_width=True, key="step4_next_btn"):
            st.session_state.step = 5
            st.rerun()


# --- TRIN 5: WATCHLIST & AKTIVERING (MED FEJLSIKRET INLINE-DOWNLOAD OG VISNING) ---
elif st.session_state.step == 5:
    st.subheader(_t("Udsendelse & Hack din portefølje", "Sändning & Aktivera", "Utsendelse & Aktiver", "Lähetys & Aktivoi", "Delivery & Activate your LLM Council"))
    
    # Trin 5 Watchlist input - Synkroniseret direkte med din Session State!
    watchlist_str = ", ".join(st.session_state.watchlist_list)
    watchlist_input = st.text_input(
        _t("Monitorer selskaber i Watchlist (kommasepareret):", "Övervaka bolag i bevakningslistan (kommaseparerat):", "Monitorer selskaper i Watchlist (kommaseparert):", "Seuraa yrityksiä tarkkailulistalla (pilkuilla erotettuna):", "Monitor tickers in your Watchlist (comma-separated):"),
        value=watchlist_str
    )
    st.session_state.watchlist_list = [t.strip().upper() for t in watchlist_input.split(",") if t.strip()]

    st.write("---")
    
    # Validation checks
    validation_passed = True
    if not st.session_state.is_new_investor and not st.session_state.investor_holdings:
        validation_passed = False
        st.error(_t(
            "⚠️ Fejl: Du skal tilføje mindst én aktiv position under Trin 4 for at kunne generere rapporten.",
            "⚠️ Fel: Du måste lägga till minst ett aktivt innehav under Steg 4 för att generera rapporten.",
            "⚠️ Feil: Du må legge til minst én aktiv posisjon under Trinn 4 for å kunne generere rapporten.",
            "⚠️ Virhe: Sinun on lisättävä vähintään yksi sijoitus Vaiheessa 4 luodaksesi raportin.",
            "⚠️ Error: You must add at least one active position under Step 4 to run your briefing."
        ))
        if st.button(_t("Gå direkte til Trin 4 for at tilføje aktiver ➔", "Gå direkt till Steg 4 för att lägga till tillgångar ➔", "Gå direkte til Trinn 4 for å legge til aktiver ➔", "Siirry suoraan Vaiheeseen 4 lisätäksesi sijoituksia ➔", "Go directly to Step 4 to add assets ➔"), use_container_width=True):
            st.session_state.step = 4
            st.rerun()
            
    elif not st.session_state.user_email or "@" not in st.session_state.user_email:
        validation_passed = False
        st.error(_t("⚠️ Fejl: Du mangler at angive en gyldig e-mailadresse i din profil under Trin 2.", "⚠️ Fel: Du måste ange en giltig e-postadress i din profil under Steg 2.", "⚠️ Feil: Du mangler å oppgje en gyldig e-postadresse i din profil under Trinn 2.", "⚠️ Virhe: Sähköpostiosoitetta ei ole annettu profiilissasi Vaiheessa 2.", "⚠️ Error: Please enter a valid email address in your profile under Step 2."))

    # Aktiveringsknapper
    col_b1, col_b2 = st.columns([2, 1])
    with col_b1:
        if st.button("🚀 " + _t("Kør LLM Council & Send min første rapport", "Kör LLM Council & Skicka min första rapport", "Kjør LLM Council & Send min første rapport", "Käynnistä LLM Council & Lähetä ensimmäinen raporttini", "Run LLM Council & Send my first report"), use_container_width=True, disabled=(not validation_passed)):
            with st.spinner(_t("Screening mod Sharia- og gældskrav samt dannelse af podcast... Det tager ca. 60 sekunder.", "Screenar mot Shariah-skuldgränser och genererar podcast... Tar cirka 60 sekunder.", "Screening mot Shariah-gjeldsgrenser og dannelse av podcast... Tar ca. 60 sekunder.", "Tarkistetaan Shariah-velkarajoja ja luodaan podcastia... Tämä kestää noin 60 sekuntia.", "Screening against Shariah debt-limits and compiling your podcast... This takes about 60 seconds.")):
                success, msg = asyncio.run(process_instant_briefing(
                    st.session_state.user_email,
                    st.session_state.investor_holdings,
                    st.session_state.watchlist_list,
                    st.session_state.targets,
                    st.session_state.user_name,
                    st.session_state.horizon,
                    st.session_state.is_new_investor,
                    []
                ))
                if success:
                    st.success(f"Udført! {msg}")
                    st.balloons()
                    st.rerun() # Genindlæser siden så download-links og lydafspilleren dukker op med det samme!
                else:
                    st.error(f"Fejl: {msg}")

    # FEJLSIKRET DOWNLOAD & AFSPILNING DIREKTE PÅ HJEMSESIDEN (HVIS GENERERET)
    if st.session_state.generated_report:
        st.write("---")
        st.subheader("📥 " + _t("Fejlsikring: Download dine filer direkte her", "Failsafe: Ladda ner dina filer direkt här", "Feilsikring: Last ned filene dine direkte her", "Varatila: Lataa tiedostosi suoraan tästä", "Failsafe: Download your files directly here"))
        st.write(_t("Hvis du ikke har modtaget e-mailen endnu, kan du downloade dine filer direkte nedenfor samt læse rapporten og lytte til podcasten med det samme:", "Om du inte har fått e-postmeddelandet än kan du ladda ner dina filer direkt nedan samt läsa rapporten och lyssna på podcasten här:", "Hvis du ikke har mottatt e-posten ennå, kan du laste ned filene direkte nedenfor samt lese rapporten og høre på podcasten her:", "Jos et ole vielä saanut sähköpostia, voit ladata tiedostosi suoraan alta sekä lukea raportin ja kuunnella podcastin heti tästä:", "If you haven't received your email yet, you can download your files directly below, as well as read the report and listen to the podcast on-screen right now:"))
        
        # Download Briefing (HTML format, som kan åbnes i Word)
        st.download_button(
            label="📄 " + _t("Hent skriftlig briefing (HTML / Word format)", "Ladda ner skriftlig briefing (HTML / Word format)", "Last ned skriftlig briefing (HTML / Word format)", "Lataa kirjallinen raportti (HTML / Word muodossa)", "Download written briefing (HTML / Word format)"),
            data=st.session_state.generated_report,
            file_name=f"{st.session_state.user_name}_Strategic_Briefing.html",
            mime="text/html",
            use_container_width=True
        )
        
        # Download Podcast (MP3)
        if st.session_state.generated_audio_bytes:
            st.download_button(
                label="🔊 " + _t("Hent lyd-podcast (MP3-format)", "Ladda ner ljud-podcast (MP3-format)", "Last ned lyd-podcast (MP3-format)", "Lataa audio-podcast (MP3-muodossa)", "Download audio podcast (MP3 format)"),
                data=st.session_state.generated_audio_bytes,
                file_name=f"{st.session_state.user_name}_LLM_Council_Podcast.mp3",
                mime="audio/mp3",
                use_container_width=True
            )
            
            # Inline afspiller
            st.write("📢 " + _t("Afspil podcast direkte på siden:", "Spela upp podcast direkt på sidan:", "Spill av podcast direkte på siden:", "Toista podcast suoraan tästä:", "Play podcast directly on this page:"))
            st.audio(st.session_state.generated_audio_bytes, format="audio/mp3")

        # Inline visning af selve rapporten
        st.write(" ")
        with st.expander("🗳️ " + _t("Læs hele rapporten på skærmen her ▾", "Läs hela rapporten på skärmen här ▾", "Les hele rapporten på skjermen her ▾", "Lue koko raportti näytöllä tästä ▾", "Read Full Report On-Screen Here ▾")):
            components.html(st.session_state.generated_report, height=500, scrolling=True)

    with col_b2:
        if st.session_state.investor_holdings and validation_passed:
            excel_bytes = generate_excel_template_bytes(
                st.session_state.investor_holdings, 
                st.session_state.watchlist_list,
                st.session_state.targets,
                {}
            )
            st.download_button(
                label="📥 " + _t("Hent mit Excel-ark", "Ladda ner mitt Excel-ark", "Download Excel-ark", "Lataa Excel-salkkuni", "Download Excel Sheet"),
                data=excel_bytes,
                file_name=f"{st.session_state.user_name}_Live_Portfolio.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="step5_excel_btn"
            )

    # Navigationsknapper
    st.write(" ")
    col_prev, col_reset = st.columns(2)
    with col_prev:
        if st.button("⬅ " + _t("Tilbage", "Tillbaka", "Tilbake", "Takaisin", "Back"), use_container_width=True, key="step5_prev_btn"):
            st.session_state.step = 4
            st.rerun()
    with col_reset:
        if st.button("🔄 " + _t("Start forfra (Trin 1)", "Starta om (Steg 1)", "Start forfra (Trinn 1)", "Aloita alusta (Vaihe 1)", "Start over (Step 1)"), use_container_width=True, key="step5_reset_btn"):
            st.session_state.step = 1
            st.rerun()


# =====================================================================
#  NATIVE, RESPONSIV DISCLAIMER
# =====================================================================
st.write(" ")
st.warning(_t(
    "Legal Disclaimer:\n\n"
    "LLM Council er et automatiseret, AI-baseret informations- og inspirationsværktøj til personligt brug. "
    "Vi tilbyder IKKE autoriseret eller licenseret finansiel rådgivning, og vi foretager ikke formelle investeringsbeslutninger på dine vegne.\n\n"
    "Finansielle markeder indebærer altid en risiko for tab, og Shariah-fortolkninger kan variere på tværs av forskellige retslærde og madhabs. "
    "Du bør altid basere dine endelige investeringsvalg på dine egne vurderinger, personlige overbevisninger og sund fornuft.\n\n"
    "For en uafhængig og manuel revision af gældsforhold, regnskabstal og compliance anbefaler vi at anvende det anerkendte værktøj Zoya Finance Platform.",
    
    "Legal Disclaimer:\n\n"
    "LLM Council är ett automatiserat, AI-baserat informations- och inspirationsverktyg för personligt bruk. "
    "Vi erbjuder INTE auktoriserad eller licensierad finansiell rådgivning, och vi fattar inte formella investeringsbeslut för din räkning.\n\n"
    "Finansiella marknader innebär alltid en risk för förlust, och Shariah-tolkningar kan variera mellan olika rättslärda och madhabs. "
    "Du bör alltid basera dina slutgiltiga investeringsval på dina egna bedömningar, personliga övertygelser och sunt förnuft.\n\n"
    "For en oberoende och manuell granskning av skuldkvoter, finansiella siffror och compliance rekommenderad att använda det erkända verktyget Zoya Finance Platform.",
    
    "Legal Disclaimer:\n\n"
    "LLM Council er et automatisert, AI-basert informasjons- og inspirasjonsverktøy til personlig bruk. "
    "Vi tilbyr IKKE autorisert eller lisensiert finansiell rådgivning, og vi foretar ikke formelle investeringsbeslutninger på dine vegne.\n\n"
    "Finansielle markeder innebærer alltid en risiko for tap, og Shariah-fortolkninger kan variere på tvers av forskjellige rettslærde og madhabs. "
    "Du bør alltid basere dine endelige investeringsvalg på dine egne vurderinger, personlige overbevisninger og sunn fornuft.\n\n"
    "For en uavhengig og manuell revisjon av gjeldsforhold, regnskapstall og compliance anbefaler vi å bruke det anerkjente verktøyet Zoya Finance Platform.",
    
    "Legal Disclaimer:\n\n"
    "LLM Council on automatisoitu, tekoälypohjainen tieto- ja inspiraatiotyökalu henkilökohtaiseen käyttöön. "
    "Emme tarjoa lisensoitua taloudellista neuvontaa emmekä tee sijoituspäätöksiä puolestasi.\n\n"
    "Rahoitusmarkkinoihin liittyy aina tappion riski, ja Shariah-tulkinnat voivat vaihdella eri lakioppineiden ja madhabien välillä. "
    "Sinun on aina perustettava lopulliset sijoituspäätöksesi omiin arvioihisi, henkilökohtaisiin vakaumuksiisi ja terveeseen järkeen.\n\n"
    "Salkun osien, velkasuhteiden ja compliance-tarkistusta varten suosittelemme virallista Zoya Finance Platform -sovellusta.",
    
    "Legal Disclaimer:\n\n"
    "The LLM Council is an automated, AI-driven informational and educational inspiration tool. It is NOT a licensed financial advisor, nor does it provide personalized investment advice or regulatory financial mandates.\n\n"
    "Financial markets carry inherent risks, and Shariah-compliance interpretations can vary across different scholars and madhabs. You must always base your final investment decisions on your own research, personal convictions, and common sense.\n\n"
    "To manually audit and double-check Shariah-compliance, financial health, or business profiles, we highly recommend utilizing the official Zoya Finance Platform."
))
